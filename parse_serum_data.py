"""
Parse the messy serum data 0507.xlsx into clean long-format CSV.
Structure: left side = 15kg/45kg wide format (metabolites x replicates),
right side = 75kg/105kg stacked individual data.
"""
import openpyxl
import pandas as pd
import numpy as np
import re

wb = openpyxl.load_workbook('phenotype/serum data 0507.xlsx', data_only=True)
ws = wb['Sheet1']

# Read all raw data
all_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    all_rows.append(list(row))

# The header row (row 0) defines metabolite columns on right side (cols V-AP, index 21-41)
right_metabolites = all_rows[0][21:42]  # Tau through Arg
print("Right-side metabolites:", right_metabolites)

records = []
flags = []  # suspicious values

# ============================================================
# Section 1: 15kg stage (rows 1-25, 0-indexed)
# ============================================================
STAGE_15_START = 1
STAGE_15_END = 25

for row_idx in range(STAGE_15_START, STAGE_15_END + 1):
    row = all_rows[row_idx]
    metabolite_15 = row[0]  # col A
    if metabolite_15 is None:
        continue

    metabolite_name = str(metabolite_15).strip()

    # DLY15 values (cols B-I = index 1-8)
    dly15_vals = []
    for c in range(1, 9):
        v = row[c]
        if v is not None:
            dly15_vals.append(float(v))
    if dly15_vals:
        records.append({
            'metabolite': metabolite_name,
            'breed': 'DLY',
            'stage_kg': 15,
            'tissue': 'serum',
            'group': 'DLY15',
            'n': len(dly15_vals),
            'mean': np.mean(dly15_vals),
            'sd': np.std(dly15_vals, ddof=1),
            'values': dly15_vals
        })

    # TFB15 values (cols L-S = index 11-18)
    tfb15_vals = []
    for c in range(11, 19):
        v = row[c]
        if v is not None:
            tfb15_vals.append(float(v))
    if tfb15_vals:
        records.append({
            'metabolite': metabolite_name,
            'breed': 'TFB',
            'stage_kg': 15,
            'tissue': 'serum',
            'group': 'TFB15',
            'n': len(tfb15_vals),
            'mean': np.mean(tfb15_vals),
            'sd': np.std(tfb15_vals, ddof=1),
            'values': tfb15_vals
        })

# ============================================================
# Section 2: 45kg stage (rows 27-51, 0-indexed)
# ============================================================
STAGE_45_START = 28
STAGE_45_END = 51

for row_idx in range(STAGE_45_START, STAGE_45_END + 1):
    row = all_rows[row_idx]
    metabolite_45_dly = row[0]   # col A — correct for DLY45
    metabolite_45_tfb = row[10]  # col K — correct for TFB45 (shifted vs col A)
    if metabolite_45_dly is None and metabolite_45_tfb is None:
        continue

    # DLY45 values (cols B-I = index 1-8), label from col A
    if metabolite_45_dly is not None:
        met_dly = str(metabolite_45_dly).strip()
        dly45_vals = [float(row[c]) for c in range(1, 9) if row[c] is not None]
        if dly45_vals:
            records.append({
                'metabolite': met_dly,
                'breed': 'DLY',
                'stage_kg': 45,
                'tissue': 'serum',
                'group': 'DLY45',
                'n': len(dly45_vals),
                'mean': np.mean(dly45_vals),
                'sd': np.std(dly45_vals, ddof=1),
                'values': dly45_vals
            })

    # TFB45 values (cols L-S = index 11-18), label from col K (offset vs col A)
    if metabolite_45_tfb is not None:
        met_tfb = str(metabolite_45_tfb).strip()
        tfb45_vals = [float(row[c]) for c in range(11, 19) if row[c] is not None]
        if tfb45_vals:
            records.append({
                'metabolite': met_tfb,
                'breed': 'TFB',
                'stage_kg': 45,
                'tissue': 'serum',
                'group': 'TFB45',
                'n': len(tfb45_vals),
                'mean': np.mean(tfb45_vals),
                'sd': np.std(tfb45_vals, ddof=1),
                'values': tfb45_vals
            })

# ============================================================
# Section 3: 75kg and 105kg on right side (rows 1-25)
# Right side: each row = one individual animal
# Col U (index 20) = group label, Cols V-AP (index 21-41) = metabolites
# ============================================================

# Map right-side group labels
for row_idx in range(STAGE_15_START, STAGE_15_END + 1):
    row = all_rows[row_idx]
    group_label = row[20]  # col U
    if group_label is None:
        continue

    group_label = str(group_label).strip()

    # Parse group: e.g. DLYM75 = DLY muscle 75kg, HZM75 = TFB muscle 75kg
    # But in serum context, these are serum samples from animals at that weight
    if 'DLY' in group_label.upper():
        breed = 'DLY'
    elif 'HZ' in group_label.upper() or 'TFB' in group_label.upper():
        breed = 'TFB'
    else:
        continue

    # Extract stage
    stage_match = re.search(r'(\d+)', group_label)
    if stage_match:
        stage = int(stage_match.group(1))
    else:
        continue

    # Get metabolite values from cols V-AP (21-41)
    values = []
    for c in range(21, 42):
        v = row[c]
        if v is not None:
            values.append(float(v))
        else:
            values.append(np.nan)

    # Record each metabolite for this individual
    for i, met_name in enumerate(right_metabolites):
        met_name = str(met_name).strip()
        if met_name and i < len(values) and not np.isnan(values[i]):
            records.append({
                'metabolite': met_name,
                'breed': breed,
                'stage_kg': stage,
                'tissue': 'serum',
                'group': group_label,
                'n': 1,
                'mean': values[i],
                'sd': np.nan,
                'values': [values[i]]
            })

# ============================================================
# Build tidy DataFrame
# ============================================================
rows_out = []
for rec in records:
    for vi, v in enumerate(rec['values']):
        rows_out.append({
            'metabolite': rec['metabolite'],
            'breed': rec['breed'],
            'stage_kg': rec['stage_kg'],
            'group': rec['group'],
            'rep': vi + 1,
            'value': v
        })

df = pd.DataFrame(rows_out)
print(f"\nTotal records: {len(df)}")
print(f"Groups: {sorted(df['group'].unique())}")
print(f"Metabolites: {sorted(df['metabolite'].unique())}")
print(f"Stages: {sorted(df['stage_kg'].unique())}")

# ============================================================
# Flag suspicious TFB45 values
# ============================================================
print("\n=== TFB45 vs other TFB stages: mean comparison ===")
pivot = df.pivot_table(
    values='value', index='metabolite', columns='stage_kg', aggfunc='mean'
)
# Filter to key AAs
key_aas = ['Arg', 'Val', 'a-AAA', 'Thr', 'Ser', 'Glu', 'Gly', 'Ala',
           'Cys', 'Met', 'Ile', 'Leu', 'Tyr', 'Phe', 'Lys', 'His',
           'Urea', 'Tau', 'Asp', 'Sar', 'b-Ala', 'Trp']
for aa in key_aas:
    if aa in pivot.index:
        vals = pivot.loc[aa]
        print(f"{aa:8s} | 15kg: {vals.get(15, np.nan):8.4f} | 45kg: {vals.get(45, np.nan):8.4f} | 75kg: {vals.get(75, np.nan):8.4f} | 105kg: {vals.get(105, np.nan):8.4f}")

# Flag anomalies: any TFB45 value > 5x the median of other TFB stages
print("\n=== SUSPICIOUS TFB45 values (|z-score| > 3 vs other stages) ===")
for met in df['metabolite'].unique():
    tfb_all = df[(df['breed'] == 'TFB') & (df['metabolite'] == met)]
    tfb45 = tfb_all[tfb_all['stage_kg'] == 45]['value']
    tfb_other = tfb_all[tfb_all['stage_kg'] != 45]['value']

    if len(tfb45) == 0 or len(tfb_other) < 3:
        continue

    other_mean = tfb_other.mean()
    other_std = tfb_other.std()
    if other_std < 1e-9:
        continue

    for _, row in tfb_all[tfb_all['stage_kg'] == 45].iterrows():
        z = (row['value'] - other_mean) / other_std
        if abs(z) > 3:
            flags.append({
                'metabolite': met,
                'group': row['group'],
                'rep': row['rep'],
                'value': row['value'],
                'z_vs_other_stages': z,
                'other_stages_mean': other_mean
            })
            print(f"  {met:8s} TFB45 rep{int(row['rep'])}: value={row['value']:.4f}, "
                  f"other_stages_mean={other_mean:.4f}, z={z:.1f}")

# Save
df.to_csv('serum_all_tidy.csv', index=False)
print(f"\nSaved serum_all_tidy.csv ({len(df)} rows)")

# Summary table (group-level means)
summary = df.groupby(['metabolite', 'breed', 'stage_kg', 'group']).agg(
    mean=('value', 'mean'),
    sd=('value', 'std'),
    n=('value', 'count')
).reset_index()
summary.to_csv('serum_summary.csv', index=False)
print(f"Saved serum_summary.csv ({len(summary)} rows)")

# Flag summary
if flags:
    flags_df = pd.DataFrame(flags)
    flags_df.to_csv('serum_tfb45_flags.csv', index=False)
    print(f"Saved serum_tfb45_flags.csv ({len(flags_df)} flagged values)")
else:
    print("No extreme outliers detected in TFB45 with current threshold.")
    print("(The reported issues with Arg=25.4 etc may be in the isotope data file, not serum.)")
