#!/usr/bin/env python3
"""Update integrated Excel with corrected isotope 45→75 kg analysis.
Rebuilds the isotope sheets with individual data statistics + Welch's two-tailed P-values.
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re

# =============================================================================
# Styling
# =============================================================================
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUBHEADER_FONT = Font(name='Arial', size=10, bold=True)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=12, bold=True)
NOTE_FONT = Font(name='Arial', size=9, italic=True)
WARN_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = THIN_BORDER

def style_row(ws, row, ncols, bold_first=False):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT if (bold_first and c==1) else BODY_FONT
        cell.alignment = CENTER if c>1 else LEFT
        cell.border = THIN_BORDER

def subheader_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEADER_FILL; cell.font = SUBHEADER_FONT
        cell.alignment = CENTER; cell.border = THIN_BORDER

def auto_width(ws, min_w=10, max_w=32):
    for col_cells in ws.columns:
        cl = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                w = sum(2 if ord(c)>127 else 1 for c in str(cell.value))
                lengths.append(w)
        if lengths:
            ws.column_dimensions[cl].width = min(max(max(lengths)+2, min_w), max_w)

def fmt_ms(mean, sem, dec=2):
    if np.isnan(mean) or np.isnan(sem): return f'{mean:.{dec}f}'
    return f'{mean:.{dec}f} \xb1 {sem:.{dec}f}'

def sig(p):
    if np.isnan(p): return ''
    if p<0.001: return '***'
    if p<0.01: return '**'
    if p<0.05: return '*'
    return 'ns'

def fmt_p(p):
    if np.isnan(p): return ''
    return f'{p:.4f} {sig(p)}'

# =============================================================================
# Raw isotope individual data
# =============================================================================

# 75 kg: [Q_protein, PS_gN, NS_abs, PS_prot, deg_abs?, PD_gN, PD_prot, dep_prot, dep_gN, ret_synth]
dl75 = {
    'DLY-2': [25.215,3.383,91.771,21.145,64.149,2.365,14.781,6.365,1.018,30.10],
    'DLY-3': [29.905,4.347,121.161,27.171,89.709,3.219,20.118,7.053,1.128,25.96],
    'DLY-4': [30.173,4.199,115.989,26.244,86.087,3.117,19.478,6.766,1.082,25.78],
    'DLY-5': [27.447,3.718,101.790,23.240,74.562,2.724,17.024,6.216,0.995,26.75],
    'DLY-6': [28.088,4.069,108.856,25.434,74.607,2.789,17.432,8.002,1.280,31.46],
    'DLY-7': [21.347,2.913,82.259,18.205,49.513,1.753,10.958,7.247,1.160,39.81],
}
tf75 = {
    'TFB-1': [20.42,2.50,66.72,15.63,49.45,1.85,11.59,4.05,0.65,25.88],
    'TFB-2': [18.99,2.17,58.16,13.58,44.29,1.65,10.34,3.24,0.52,23.85],
    'TFB-3': [16.40,1.95,50.42,12.20,31.62,1.22,7.65,4.55,0.73,37.30],
    'TFB-5': [19.39,2.44,64.05,15.27,43.44,1.66,10.36,4.91,0.79,32.17],
    'TFB-6': [17.87,2.17,57.81,13.55,39.47,1.48,9.75,3.80,0.69,31.74],
    'TFB-7': [18.61,2.25,59.43,14.04,41.65,1.57,9.94,4.11,0.67,30.19],
}

# 45 kg: [Q_gN, Turnover_prot, PS_gN, NS_abs, PS_prot, D_abs, PD_gN, PD_prot, dep_prot, dep_N, ret_synth]
d45 = {
    'DLY45-1': [5.576,34.853,5.03,95.19,31.41,68.39,3.61,22.56,8.84,1.41,28.16],
    'DLY45-2': [5.786,36.165,5.40,103.88,33.77,71.33,3.71,23.19,10.58,1.69,31.34],
    'DLY45-3': [5.303,33.146,4.86,92.13,30.40,61.00,3.22,20.12,10.27,1.64,33.79],
    'DLY45-4': [5.578,34.863,5.15,101.23,32.21,70.54,3.59,22.45,9.76,1.56,30.31],
    'DLY45-5': [7.393,46.208,7.08,132.15,44.26,101.61,5.44,34.03,10.23,1.64,23.11],
    'DLY45-6': [8.923,55.770,8.57,164.81,53.58,130.75,6.80,42.51,11.08,1.77,20.67],
    'DLY45-7': [4.429,27.682,3.92,76.50,24.51,47.80,2.45,15.32,9.20,1.47,37.51],
    'DLY45-8': [6.794,42.460,6.33,119.04,39.57,90.05,4.79,29.93,9.64,1.54,24.36],
}
t45 = {
    'TFB45-1': [6.790,42.43,6.12,108.98,38.25,84.98,4.77,29.83,8.42,1.35,22.02],
    'TFB45-2': [6.945,43.41,5.92,103.74,37.01,87.30,4.98,31.14,5.87,0.94,15.85],
    'TFB45-3': [6.641,41.51,5.95,108.48,37.18,85.19,4.67,29.20,7.98,1.28,21.48],
    'TFB45-4': [6.814,42.59,5.49,100.55,34.30,87.82,4.79,29.96,4.34,0.69,12.66],
    'TFB45-5': [7.198,44.99,6.20,112.28,38.72,93.64,5.17,32.30,6.43,1.03,16.60],
    'TFB45-6': [7.378,46.11,6.28,112.27,39.28,94.11,5.27,32.93,6.35,1.02,16.17],
    'TFB45-7': [7.624,47.65,6.87,123.19,42.96,99.38,5.55,34.66,8.30,1.33,19.33],
    'TFB45-8': [8.176,51.10,7.39,129.95,46.21,106.91,6.08,38.01,8.19,1.31,17.73],
}

# Extract individual series for 45 & 75 kg
iso_series = {}  # iso_series[stage][param][breed] = np.array

param_45_idx = {'Q':0, 'PS':2, 'PD':6, 'dep':9, 'RS':10}
param_75_idx = {'Q':0, 'PS':1, 'PD':5, 'dep':8, 'RS':9}

iso_series['45 kg'] = {}
iso_series['75 kg'] = {}

for p, idx45, idx75 in [('Q',0,0), ('PS',2,1), ('PD',6,5), ('dep',9,8), ('RS',10,9)]:
    iso_series['45 kg'][p] = {
        'DLY': np.array([v[idx45] for v in d45.values()]),
        'TFB': np.array([v[idx45] for v in t45.values()]),
    }
    if p == 'Q':
        # 75 kg: Q_gN = col0/6.25
        iso_series['75 kg'][p] = {
            'DLY': np.array([v[0]/6.25 for v in dl75.values()]),
            'TFB': np.array([v[0]/6.25 for v in tf75.values()]),
        }
    else:
        iso_series['75 kg'][p] = {
            'DLY': np.array([v[idx75] for v in dl75.values()]),
            'TFB': np.array([v[idx75] for v in tf75.values()]),
        }

# Compute descriptive stats
iso_desc = {}
for stage in ['45 kg', '75 kg']:
    iso_desc[stage] = {}
    for p in ['Q','PS','PD','dep','RS']:
        iso_desc[stage][p] = {}
        for breed in ['DLY','TFB']:
            v = iso_series[stage][p][breed]
            iso_desc[stage][p][breed] = {
                'n': len(v), 'mean': np.mean(v), 'sd': np.std(v,ddof=1),
                'sem': np.std(v,ddof=1)/np.sqrt(len(v))
            }

# Breed comparison (Welch two-tailed)
iso_breed_p = {}
for stage in ['45 kg', '75 kg']:
    iso_breed_p[stage] = {}
    for p in ['Q','PS','PD','dep','RS']:
        dly = iso_series[stage][p]['DLY']
        tfb = iso_series[stage][p]['TFB']
        t, pval = stats.ttest_ind(dly, tfb, equal_var=False)
        pooled_sd = np.sqrt((np.std(dly,ddof=1)**2 + np.std(tfb,ddof=1)**2)/2)
        d = (np.mean(dly)-np.mean(tfb))/pooled_sd if pooled_sd>0 else np.nan
        iso_breed_p[stage][p] = {'t':t, 'p':pval, 'cohens_d':d}

# Adjacent stage comparison (independent t-test)
iso_adj_p = {}
for breed in ['DLY','TFB']:
    iso_adj_p[breed] = {}
    for p in ['Q','PS','PD','dep','RS']:
        v45 = iso_series['45 kg'][p][breed]
        v75 = iso_series['75 kg'][p][breed]
        t, pval = stats.ttest_ind(v45, v75, equal_var=False)
        iso_adj_p[breed][p] = {'t':t, 'p':pval}

# =============================================================================
# Read growth performance data (from previous integrated build)
# =============================================================================

growth = pd.read_csv('growth_performance_tidy.csv')
breeds = ['DLY', 'TFB']
stages_4 = ['15 kg', '45 kg', '75 kg', '105 kg']
adj_pairs = [('15 kg','45 kg'),('45 kg','75 kg'),('75 kg','105 kg')]

gp_vars = {'ADG_kg_d':'ADG, kg/d', 'ADFI_kg_d':'ADFI, kg/d', 'F_G':'F/G'}

gp_desc = {}
for var in gp_vars:
    gp_desc[var] = {}
    for breed in breeds:
        gp_desc[var][breed] = {}
        for stage in stages_4:
            d = growth[(growth['Breed']==breed)&(growth['Stage']==stage)][var]
            gp_desc[var][breed][stage] = {
                'n':len(d), 'mean':d.mean(), 'sd':d.std(), 'sem':d.sem()
            }

gp_breed_p = {}
for var in gp_vars:
    gp_breed_p[var] = {}
    for stage in stages_4:
        d1 = growth[(growth['Breed']=='DLY')&(growth['Stage']==stage)][var]
        d2 = growth[(growth['Breed']=='TFB')&(growth['Stage']==stage)][var]
        t,p = stats.ttest_ind(d1, d2, equal_var=False)
        psd = np.sqrt((d1.std()**2+d2.std()**2)/2)
        d_cohen = (d1.mean()-d2.mean())/psd if psd>0 else np.nan
        gp_breed_p[var][stage] = {'t':t,'p':p,'cohens_d':d_cohen}

gp_adj_p = {}
for var in gp_vars:
    gp_adj_p[var] = {}
    for breed in breeds:
        gp_adj_p[var][breed] = {}
        for s1,s2 in adj_pairs:
            d1 = growth[(growth['Breed']==breed)&(growth['Stage']==s1)][var]
            d2 = growth[(growth['Breed']==breed)&(growth['Stage']==s2)][var]
            t,p = stats.ttest_ind(d1, d2, equal_var=False)
            gp_adj_p[var][breed][f'{s1} vs {s2}']={'t':t,'p':p}

# =============================================================================
# Read N balance & isotope source data for 15/105 kg
# =============================================================================

src = pd.read_excel('phenotype/data nb isotope.xlsx', sheet_name='Sheet2', header=None)

def parse_mean_sd(v):
    if pd.isna(v) or str(v).strip()=='': return (np.nan, np.nan)
    s = str(v).strip()
    if s.startswith('<') or s.startswith('>'): return (np.nan, np.nan)
    for sep in ['±', '\xb1']:
        if sep in s:
            parts = s.split(sep)
            try: return (float(parts[0].strip()), float(parts[1].strip()))
            except: pass
    try: return (float(s), np.nan)
    except: return (np.nan, np.nan)

def parse_pv(v):
    if pd.isna(v) or str(v).strip()=='': return (np.nan, '')
    s = str(v).strip()
    if s.startswith('<'):
        nums = re.findall(r'[\d.]+', s)
        return (float(nums[0]) if nums else np.nan, s)
    try: return (float(s), s)
    except: return (np.nan, s)

# Parse N balance source for 15 & 105 kg
nb_items = ['N intake, g/d','FN, g/d','UN, g/d','TN, g/d','RN, g/d','N retention, %','N ABV, %']
nb_labels = {
    'N intake, g/d':'N intake, g/d', 'FN, g/d':'Fecal N, g/d',
    'UN, g/d':'Urinary N, g/d', 'TN, g/d':'Total N excretion, g/d',
    'RN, g/d':'Retained N, g/d', 'N retention, %':'N retention rate, %',
    'N ABV, %':'N ABV, %',
}

mean_cols = [
    ('DLY','15 kg',1),('TFB','15 kg',2),('DLY','45 kg',4),('TFB','45 kg',5),
    ('DLY','75 kg',7),('TFB','75 kg',8),('DLY','105 kg',10),('TFB','105 kg',11),
]
breed_p_cols = [('15 kg',3),('45 kg',6),('75 kg',9),('105 kg',12)]
adj_p_cols = [
    ('DLY','15 kg vs 45 kg',14),('DLY','45 kg vs 75 kg',15),('DLY','75 kg vs 105 kg',16),
    ('TFB','15 kg vs 45 kg',17),('TFB','45 kg vs 75 kg',18),('TFB','75 kg vs 105 kg',19),
]

def parse_nb():
    data = {}
    for i, item in enumerate(nb_items):
        row = src.iloc[1+i]
        data[item] = {}
        for breed,stage,col in mean_cols:
            m,sd = parse_mean_sd(row[col])
            data[item][f'{breed}_{stage}'] = {'mean':m,'sd':sd}
        for stage,col in breed_p_cols:
            pv,_ = parse_pv(row[col])
            data[item][f'P_breed_{stage}'] = pv
        for breed,pair,col in adj_p_cols:
            pv,_ = parse_pv(row[col])
            data[item][f'P_{breed}_{pair}'] = pv
    return data

nb_data = parse_nb()

# Parse isotope source for 15 & 105 kg
iso_items = [
    'N flux, g/kg BW^0.75/d','Protein synthesis, N g/kg BW^0.75/d',
    'Protein degradation, N g/kg BW^0.75/d','Protein deposition, N g/kg BW^0.75/d',
    'Protein retention/synthesis, %',
]
iso_short = {
    'N flux, g/kg BW^0.75/d':'N flux',
    'Protein synthesis, N g/kg BW^0.75/d':'Protein synthesis',
    'Protein degradation, N g/kg BW^0.75/d':'Protein degradation',
    'Protein deposition, N g/kg BW^0.75/d':'Protein deposition',
    'Protein retention/synthesis, %':'Retention/Synthesis',
}

def parse_iso():
    data = {}
    for i, item in enumerate(iso_items):
        row = src.iloc[10+i]
        data[item] = {}
        for breed,stage,col in mean_cols:
            m,sd = parse_mean_sd(row[col])
            data[item][f'{breed}_{stage}'] = {'mean':m,'sd':sd}
        for stage,col in breed_p_cols:
            pv,_ = parse_pv(row[col])
            data[item][f'P_breed_{stage}'] = pv
        for breed,pair,col in adj_p_cols:
            pv,_ = parse_pv(row[col])
            data[item][f'P_{breed}_{pair}'] = pv
    return data

iso_src = parse_iso()

# =============================================================================
# Build Excel workbook
# =============================================================================

wb = Workbook()

# ===== Sheet 1: Growth Performance Descriptive =====
ws = wb.active
ws.title = '1_Growth_Descriptive'
ncols = 9

ws.merge_cells('A1:I1')
ws.cell(row=1, column=1, value='Growth Performance — Descriptive Statistics (Mean ± SEM)').font = TITLE_FONT

ws.cell(row=3, column=1, value='Item')
col = 2
for stage in stages_4:
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.cell(row=3, column=col, value=stage)
    col += 2
style_header(ws, 3, ncols)

ws.cell(row=4, column=1, value='Breed')
for si in range(4):
    for bi, breed in enumerate(breeds):
        ws.cell(row=4, column=2+si*2+bi, value=breed)
subheader_row(ws, 4, ncols)

for i, (var, label) in enumerate(gp_vars.items()):
    r = 5+i
    ws.cell(row=r, column=1, value=label)
    for si, stage in enumerate(stages_4):
        for bi, breed in enumerate(breeds):
            d = gp_desc[var][breed][stage]
            ws.cell(row=r, column=2+si*2+bi, value=fmt_ms(d['mean'],d['sem'],dec=3))
    style_row(ws, r, ncols)

r_n = 8
ws.cell(row=r_n, column=1, value='n')
for si, stage in enumerate(stages_4):
    for bi, breed in enumerate(breeds):
        ws.cell(row=r_n, column=2+si*2+bi, value=gp_desc['ADG_kg_d'][breed][stage]['n'])
style_row(ws, r_n, ncols, bold_first=True)

r_note = 9
ws.merge_cells(start_row=r_note, start_column=1, end_row=r_note, end_column=ncols)
ws.cell(row=r_note, column=1, value='Mean ± SEM. Breed comparison: Welch\'s independent t-test (two-tailed).').font = NOTE_FONT
auto_width(ws)

# ===== Sheet 2: Growth Breed P =====
ws = wb.create_sheet('2_Growth_Breed_P')
ncols = 7
ws.merge_cells('A1:G1')
ws.cell(row=1, column=1, value='Growth Performance — Breed Comparison (DLY vs TFB)').font = TITLE_FONT

for ci, h in enumerate(['Item','15 kg','45 kg','75 kg','105 kg','Test','Note'], 1):
    ws.cell(row=3, column=ci, value=h)
style_header(ws, 3, ncols)

for i, (var, label) in enumerate(gp_vars.items()):
    r = 4+i
    ws.cell(row=r, column=1, value=label)
    for j, stage in enumerate(stages_4):
        bp = gp_breed_p[var][stage]
        txt = f'P = {fmt_p(bp["p"])}\nd = {bp["cohens_d"]:.3f}'
        ws.cell(row=r, column=2+j, value=txt)
    ws.cell(row=r, column=6, value="Welch's t")
    ws.cell(row=r, column=7, value='')
    style_row(ws, r, ncols)
auto_width(ws)

# ===== Sheet 3: Growth Adjacent P =====
ws = wb.create_sheet('3_Growth_Adjacent_P')
ncols = 7
ws.merge_cells('A1:G1')
ws.cell(row=1, column=1, value='Growth Performance — Adjacent Stage Comparison (Within Breed)').font = TITLE_FONT

for ci, h in enumerate(['Item','Breed','15 vs 45 kg','45 vs 75 kg','75 vs 105 kg','Test','Note'], 1):
    ws.cell(row=3, column=ci, value=h)
style_header(ws, 3, ncols)

r = 4
for var, label in gp_vars.items():
    for breed in breeds:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=breed)
        for j, (s1,s2) in enumerate(adj_pairs):
            pv = gp_adj_p[var][breed][f'{s1} vs {s2}']['p']
            ws.cell(row=r, column=3+j, value=f'P = {fmt_p(pv)}')
        ws.cell(row=r, column=6, value='Independent t')
        ws.cell(row=r, column=7, value='Same breed, different pigs at different stages')
        style_row(ws, r, ncols)
        r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
ws.cell(row=r, column=1, value='Note: Independent t-test used (pig IDs not in tidy data). Original design is longitudinal — paired t-test recommended with IDs.').font = NOTE_FONT
auto_width(ws)

# ===== Sheet 4: N Balance Descriptive =====
ws = wb.create_sheet('4_NBalance_Descriptive')
ncols = 9
ws.merge_cells('A1:I1')
ws.cell(row=1, column=1, value='Nitrogen Balance — Descriptive Statistics (Mean ± SD from Source Summary)').font = TITLE_FONT

ws.cell(row=3, column=1, value='Item')
col = 2
for stage in stages_4:
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.cell(row=3, column=col, value=stage)
    col += 2
style_header(ws, 3, ncols)

ws.cell(row=4, column=1, value='Breed')
for si in range(4):
    for bi, breed in enumerate(breeds):
        ws.cell(row=4, column=2+si*2+bi, value=breed)
subheader_row(ws, 4, ncols)

for i, item in enumerate(nb_items):
    r = 5+i
    ws.cell(row=r, column=1, value=nb_labels[item])
    for si, stage in enumerate(stages_4):
        for bi, breed in enumerate(breeds):
            d = nb_data[item][f'{breed}_{stage}']
            ws.cell(row=r, column=2+si*2+bi, value=fmt_ms(d['mean'],d['sd'],dec=2))
    style_row(ws, r, ncols)

r_note = 5+len(nb_items)+1
ws.merge_cells(start_row=r_note, start_column=1, end_row=r_note, end_column=ncols)
ws.cell(row=r_note, column=1,
        value='Note: Mean ± SD from source summary file. 45 & 75 kg have raw individual data available for N balance; '
        '15 & 105 kg individual data still needed for SEM and corrected P-values.').font = NOTE_FONT
auto_width(ws)

# ===== Sheet 5: N Balance Breed P =====
ws = wb.create_sheet('5_NBalance_Breed_P')
ncols = 7
ws.merge_cells('A1:G1')
ws.cell(row=1, column=1, value='N Balance — Breed Comparison (DLY vs TFB)').font = TITLE_FONT

for ci, h in enumerate(['Item','15 kg','45 kg','75 kg','105 kg','Test','Note'], 1):
    ws.cell(row=3, column=ci, value=h)
style_header(ws, 3, ncols)

for i, item in enumerate(nb_items):
    r = 4+i
    ws.cell(row=r, column=1, value=nb_labels[item])
    for j, stage in enumerate(stages_4):
        pv = nb_data[item][f'P_breed_{stage}']
        ws.cell(row=r, column=2+j, value=fmt_p(pv))
    ws.cell(row=r, column=6, value='See source')
    style_row(ws, r, ncols)
auto_width(ws)

# ===== Sheet 6: N Balance Adjacent P =====
ws = wb.create_sheet('6_NBalance_Adjacent_P')
ncols = 7
ws.merge_cells('A1:G1')
ws.cell(row=1, column=1, value='N Balance — Adjacent Stage Comparison (Within Breed)').font = TITLE_FONT

for ci, h in enumerate(['Item','Breed','15 vs 45 kg','45 vs 75 kg','75 vs 105 kg','Test','Note'], 1):
    ws.cell(row=3, column=ci, value=h)
style_header(ws, 3, ncols)

r = 4
for item in nb_items:
    for breed in breeds:
        ws.cell(row=r, column=1, value=nb_labels[item])
        ws.cell(row=r, column=2, value=breed)
        for j, (s1,s2) in enumerate(adj_pairs):
            key = f'P_{breed}_{s1} vs {s2}'
            pv = nb_data[item].get(key, np.nan)
            ws.cell(row=r, column=3+j, value=fmt_p(pv))
        ws.cell(row=r, column=6, value='From source')
        style_row(ws, r, ncols)
        r += 1
auto_width(ws)

# ===== Sheet 7: Isotope Descriptive (CORRECTED with individual data for 45 & 75 kg) =====
ws = wb.create_sheet('7_Isotope_Descriptive')
ncols = 9
ws.merge_cells('A1:I1')
ws.cell(row=1, column=1, value='Isotope Tracer — Descriptive Statistics').font = TITLE_FONT

ws.cell(row=3, column=1, value='Item')
col = 2
for stage in stages_4:
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.cell(row=3, column=col, value=stage)
    col += 2
style_header(ws, 3, ncols)

ws.cell(row=4, column=1, value='Breed')
for si in range(4):
    for bi, breed in enumerate(breeds):
        ws.cell(row=4, column=2+si*2+bi, value=breed)
subheader_row(ws, 4, ncols)

iso_param_keys = ['N flux, g/kg BW^0.75/d','Protein synthesis, N g/kg BW^0.75/d',
                  'Protein degradation, N g/kg BW^0.75/d','Protein deposition, N g/kg BW^0.75/d',
                  'Protein retention/synthesis, %']
# Map to short keys for corrected data
iso_short_keys = ['Q','PS','PD','dep','RS']

for i, (item, sk) in enumerate(zip(iso_param_keys, iso_short_keys)):
    r = 5+i
    ws.cell(row=r, column=1, value=item)
    for si, stage in enumerate(stages_4):
        for bi, breed in enumerate(breeds):
            if stage in ['45 kg', '75 kg']:
                # Use corrected individual data
                d = iso_desc[stage][sk][breed]
                val = fmt_ms(d['mean'], d['sem'], dec=2)
            else:
                # Use source summary for 15 & 105 kg
                d = iso_src[item][f'{breed}_{stage}']
                val = fmt_ms(d['mean'], d['sd'], dec=2)
            ws.cell(row=r, column=2+si*2+bi, value=val)
    style_row(ws, r, ncols)

# n row
r_n = 5+len(iso_param_keys)
ws.cell(row=r_n, column=1, value='n (45/75 kg)')
for si, stage in enumerate(stages_4):
    if stage in ['45 kg', '75 kg']:
        for bi, breed in enumerate(breeds):
            ws.cell(row=r_n, column=2+si*2+bi, value=iso_desc[stage]['Q'][breed]['n'])
    else:
        ws.cell(row=r_n, column=2+si*2, value='(source)')
        ws.cell(row=r_n, column=3+si*2, value='(source)')
style_row(ws, r_n, ncols, bold_first=True)

# Data source note
r_src = r_n+2
ws.merge_cells(start_row=r_src, start_column=1, end_row=r_src, end_column=ncols)
ws.cell(row=r_src, column=1, value='45 & 75 kg: Mean ± SEM from individual animal data (DLY45 n=8, TFB45 n=8, DLY75 n=6, TFB75 n=6). '
         '15 & 105 kg: Mean ± SD from source summary. Bold n reflects individual data availability.').font = NOTE_FONT
auto_width(ws)

# ===== Sheet 8: Isotope Breed P (CORRECTED) =====
ws = wb.create_sheet('8_Isotope_Breed_P')
ncols = 9
ws.merge_cells('A1:I1')
ws.cell(row=1, column=1, value='Isotope Tracer — Breed Comparison (DLY vs TFB) | Welch\'s Two-tailed t-test').font = TITLE_FONT

for ci, h in enumerate(['Item','15 kg','45 kg','75 kg','105 kg','Test','Note','45kg Detail','75kg Detail'], 1):
    ws.cell(row=3, column=ci, value=h)
style_header(ws, 3, ncols)

for i, (item, sk) in enumerate(zip(iso_param_keys, iso_short_keys)):
    r = 4+i
    ws.cell(row=r, column=1, value=iso_short[item])
    # 15 & 105 kg: use source P
    for j, stage in enumerate(['15 kg','45 kg','75 kg','105 kg']):
        sname = stage
        col_idx = 2+j  # columns 2,3,4,5 (column 1 is Item name)
        if sname in ['45 kg','75 kg']:
            bp = iso_breed_p[sname][sk]
            pv = bp['p']
            ws.cell(row=r, column=col_idx, value=fmt_p(pv))
        else:
            pv = iso_src[item][f'P_breed_{sname}']
            ws.cell(row=r, column=col_idx, value=fmt_p(pv))

    ws.cell(row=r, column=6, value="Welch's t (45/75) | Source (15/105)")

    # Detail columns
    if sk in iso_breed_p['45 kg']:
        bp45 = iso_breed_p['45 kg'][sk]
        ws.cell(row=r, column=8, value=f"d={bp45['cohens_d']:.3f}, n=8,8")
        bp75 = iso_breed_p['75 kg'][sk]
        ws.cell(row=r, column=9, value=f"d={bp75['cohens_d']:.3f}, n=6,6")

    style_row(ws, r, ncols)
auto_width(ws)

# ===== Sheet 9: Isotope Adjacent P (CORRECTED for 45→75) =====
ws = wb.create_sheet('9_Isotope_Adjacent_P')
ncols = 7
ws.merge_cells('A1:G1')
ws.cell(row=1, column=1, value='Isotope Tracer — Adjacent Stage Comparison (Within Breed) | Independent t-test').font = TITLE_FONT

for ci, h in enumerate(['Item','Breed','15 vs 45 kg','45 vs 75 kg','75 vs 105 kg','Test','Note'], 1):
    ws.cell(row=3, column=ci, value=h)
style_header(ws, 3, ncols)

r = 4
for item, sk in zip(iso_param_keys, iso_short_keys):
    for breed in breeds:
        ws.cell(row=r, column=1, value=iso_short[item])
        ws.cell(row=r, column=2, value=breed)

        # 15 vs 45: from source
        pv = iso_src[item].get(f'P_{breed}_15 kg vs 45 kg', np.nan)
        ws.cell(row=r, column=3, value=fmt_p(pv))

        # 45 vs 75: CORRECTED from individual data
        if sk in iso_adj_p[breed]:
            pv_corrected = iso_adj_p[breed][sk]['p']
            ws.cell(row=r, column=4, value=fmt_p(pv_corrected))
        else:
            pv = iso_src[item].get(f'P_{breed}_45 kg vs 75 kg', np.nan)
            ws.cell(row=r, column=4, value=fmt_p(pv))

        # 75 vs 105: from source
        pv = iso_src[item].get(f'P_{breed}_75 kg vs 105 kg', np.nan)
        ws.cell(row=r, column=5, value=fmt_p(pv))

        ws.cell(row=r, column=6, value='Indep. t (45→75 corrected)')
        style_row(ws, r, ncols)
        r += 1
auto_width(ws)

# ===== Sheet 10: P-value Summary =====
ws = wb.create_sheet('10_Pvalue_Summary')
ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value='Complete P-value Matrix — All Phenotypes, All Stages').font = TITLE_FONT

# Part A: Breed comparison
r = 3
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='A. BREED COMPARISON: DLY vs TFB at Each Stage').font = BOLD_FONT
r += 1

for ci, h in enumerate(['Item','Type','15 kg','45 kg','75 kg','105 kg','Test','Data Source'], 1):
    ws.cell(row=r, column=ci, value=h)
style_header(ws, r, 8)
r += 1

# Growth
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='Growth Performance (Welch\'s t, two-tailed)')
subheader_row(ws, r, 8)
r += 1
for var, label in gp_vars.items():
    ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value='Breed P')
    for j, stage in enumerate(stages_4):
        ws.cell(row=r, column=3+j, value=fmt_p(gp_breed_p[var][stage]['p']))
    ws.cell(row=r, column=7, value="Welch's t")
    ws.cell(row=r, column=8, value='Individual data')
    style_row(ws, r, 8); r += 1

# N Balance
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='N Balance (source P-values)')
subheader_row(ws, r, 8)
r += 1
for item in nb_items:
    ws.cell(row=r, column=1, value=nb_labels[item]); ws.cell(row=r, column=2, value='Breed P')
    for j, stage in enumerate(stages_4):
        ws.cell(row=r, column=3+j, value=fmt_p(nb_data[item][f'P_breed_{stage}']))
    ws.cell(row=r, column=7, value='Source')
    ws.cell(row=r, column=8, value='Summary file')
    style_row(ws, r, 8); r += 1

# Isotope
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='Isotope Tracer (45 & 75 kg: Welch\'s t from individual data; 15 & 105 kg: source)')
subheader_row(ws, r, 8)
r += 1
for item, sk in zip(iso_param_keys, iso_short_keys):
    ws.cell(row=r, column=1, value=iso_short[item]); ws.cell(row=r, column=2, value='Breed P')
    for j, stage in enumerate(stages_4):
        if stage in ['45 kg','75 kg']:
            pv = iso_breed_p[stage][sk]['p']
        else:
            pv = iso_src[item][f'P_breed_{stage}']
        ws.cell(row=r, column=3+j, value=fmt_p(pv))
    ws.cell(row=r, column=7, value="Welch's t (45/75)")
    ws.cell(row=r, column=8, value='Individual data (45/75)')
    style_row(ws, r, 8); r += 1

# Part B: Adjacent stage comparison
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='B. ADJACENT STAGE COMPARISON: Within Each Breed').font = BOLD_FONT
r += 1

for ci, h in enumerate(['Item','Breed','15 vs 45 kg','45 vs 75 kg','75 vs 105 kg','Test','Data Source'], 1):
    ws.cell(row=r, column=ci, value=h)
style_header(ws, r, 7)
r += 1

# Growth
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(row=r, column=1, value='Growth Performance')
subheader_row(ws, r, 7)
r += 1
for var, label in gp_vars.items():
    for breed in breeds:
        ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=breed)
        for j, (s1,s2) in enumerate(adj_pairs):
            pv = gp_adj_p[var][breed][f'{s1} vs {s2}']['p']
            ws.cell(row=r, column=3+j, value=fmt_p(pv))
        ws.cell(row=r, column=6, value='Indep. t')
        style_row(ws, r, 7); r += 1

# N Balance
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(row=r, column=1, value='N Balance (source)')
subheader_row(ws, r, 7)
r += 1
for item in nb_items:
    for breed in breeds:
        ws.cell(row=r, column=1, value=nb_labels[item]); ws.cell(row=r, column=2, value=breed)
        for j, (s1,s2) in enumerate(adj_pairs):
            pv = nb_data[item].get(f'P_{breed}_{s1} vs {s2}', np.nan)
            ws.cell(row=r, column=3+j, value=fmt_p(pv))
        ws.cell(row=r, column=6, value='Source')
        style_row(ws, r, 7); r += 1

# Isotope (corrected 45→75)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(row=r, column=1, value='Isotope Tracer (45→75 corrected from individual data)')
subheader_row(ws, r, 7)
r += 1
for item, sk in zip(iso_param_keys, iso_short_keys):
    for breed in breeds:
        ws.cell(row=r, column=1, value=iso_short[item]); ws.cell(row=r, column=2, value=breed)
        # 15→45: source
        pv = iso_src[item].get(f'P_{breed}_15 kg vs 45 kg', np.nan)
        ws.cell(row=r, column=3, value=fmt_p(pv))
        # 45→75: CORRECTED
        if sk in iso_adj_p[breed]:
            pv = iso_adj_p[breed][sk]['p']
        else:
            pv = iso_src[item].get(f'P_{breed}_45 kg vs 75 kg', np.nan)
        ws.cell(row=r, column=4, value=fmt_p(pv))
        # 75→105: source
        pv = iso_src[item].get(f'P_{breed}_75 kg vs 105 kg', np.nan)
        ws.cell(row=r, column=5, value=fmt_p(pv))
        ws.cell(row=r, column=6, value='Indep. t (45→75)')
        style_row(ws, r, 7); r += 1

auto_width(ws)

# ===== Sheet 11: Statistical Method Notes =====
ws = wb.create_sheet('11_Method_Notes')

ws.merge_cells('A1:C1')
ws.cell(row=1, column=1, value='Statistical Methods & Data Sources').font = TITLE_FONT

notes = [
    ('Test type', 'Breed comparison (DLY vs TFB)', "Welch's independent t-test (two-tailed). Does not assume equal variance."),
    ('Test type', 'Adjacent stage comparison (within breed)', 'Independent t-test. Different pigs at different metabolic stages for N balance/isotope. Growth performance may use paired t-test if pig IDs available.'),
    ('Growth Performance', 'Data source', 'Individual animal data (n=22 DLY, n=8 TFB per stage). All P-values computed from raw data.'),
    ('N Balance', 'Data source', 'Source summary file (Mean±SD). 15 & 105 kg individual data not yet available. P-values from source.'),
    ('Isotope Tracer', 'Data source (15 & 105 kg)', 'Source summary file (Mean±SD). Individual data not yet available.'),
    ('Isotope Tracer', 'Data source (45 & 75 kg)', 'Individual animal data. DLY45 n=8, TFB45 n=8, DLY75 n=6, TFB75 n=6. All P-values: Welch\'s two-tailed t-test.'),
    ('Isotope Tracer', '75 kg Q (N flux)', 'Q_protein column / 6.25. First data column in raw table = protein turnover in g protein/kg^0.75/d.'),
    ('Isotope Tracer', 'Key finding 45 kg', 'DLY variance 7-9x larger than TFB at 45 kg for Q/PS/PD. Source file P-values for Q (0.045) and PD (0.044) appear to use ONE-TAILED tests — correct two-tailed Welch P = 0.102 and 0.100 respectively (both ns).'),
    ('Isotope Tracer', 'Key finding 75 kg', 'All breed comparisons highly significant (Q/PS/PD/dep P≤0.003) except Retention/Synthesis (P=0.944). TFB shows dramatic metabolic downshift from 45→75 kg (all P<0.001).'),
    ('Significance notation', '*** P<0.001, ** P<0.01, * P<0.05, ns P≥0.05', 'All P-values displayed as numbers with significance markers.'),
]

for i, (cat, topic, detail) in enumerate(notes):
    r = 3+i
    ws.cell(row=r, column=1, value=cat)
    ws.cell(row=r, column=2, value=topic)
    ws.cell(row=r, column=3, value=detail)
    style_row(ws, r, 3)
    ws.cell(row=r, column=1).font = BOLD_FONT

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 80

# =============================================================================
# Save
# =============================================================================

output_path = 'integrated_15_105kg_all_phenotypes.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
print(f'Sheets ({len(wb.sheetnames)}):')
for s in wb.sheetnames:
    print(f'  - {s}')
print('\nKey corrections applied:')
print('  ✓ 75 kg Q (N flux) identified as col0/6.25 (protein turnover per kg^0.75)')
print('  ✓ All isotope 45 & 75 kg P-values: Welch\'s two-tailed t-test from individual data')
print('  ✓ 45 kg Q and PD: correctly NS (source file used one-tailed tests)')
print('  ✓ All means ± SEM from individual data for 45 & 75 kg isotope')
