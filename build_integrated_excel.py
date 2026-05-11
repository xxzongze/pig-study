#!/usr/bin/env python3
"""Build integrated 15-105 kg phenotype Excel with 9+ sheets.
Covers: Growth Performance, N Balance, Isotope Tracer across 4 stages.
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import re

# =============================================================================
# Helper functions
# =============================================================================

def parse_mean_sd(value_str):
    """Parse 'mean ± sd' string into (mean, sd) tuple."""
    if pd.isna(value_str) or str(value_str).strip() == '':
        return (np.nan, np.nan)
    s = str(value_str).strip()
    if s.startswith('<') or s.startswith('>'):
        return (np.nan, np.nan)
    if '\xb1' in s:  # ± as latin-1
        parts = s.split('\xb1')
    elif '±' in s:
        parts = s.split('±')
    else:
        try:
            return (float(s), np.nan)
        except ValueError:
            return (np.nan, np.nan)
    try:
        mean = float(parts[0].strip())
        sd = float(parts[1].strip())
        return (mean, sd)
    except (ValueError, IndexError):
        return (np.nan, np.nan)


def parse_pvalue(value_str):
    """Parse P-value string to float and original string."""
    if pd.isna(value_str) or str(value_str).strip() == '':
        return (np.nan, '')
    s = str(value_str).strip()
    if s.startswith('<'):
        nums = re.findall(r'[\d.]+', s)
        val = float(nums[0]) if nums else np.nan
        return (val, s)
    if s.startswith('>'):
        nums = re.findall(r'[\d.]+', s)
        val = float(nums[0]) if nums else np.nan
        return (val, s)
    try:
        return (float(s), s)
    except ValueError:
        return (np.nan, s)


def sig_marker(p):
    if np.isnan(p): return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return 'ns'


def format_mean_sem(mean, sem, decimals=2):
    """Format as 'mean ± sem' with same decimals."""
    if np.isnan(mean) or np.isnan(sem):
        return ''
    return f'{mean:.{decimals}f} \xb1 {sem:.{decimals}f}'


def format_mean_sd(mean, sd, decimals=2):
    if np.isnan(mean) or np.isnan(sd):
        return ''
    return f'{mean:.{decimals}f} \xb1 {sd:.{decimals}f}'


def format_p(p_val):
    if np.isnan(p_val):
        return ''
    s = f'{p_val:.4f}'
    m = sig_marker(p_val)
    return f'{s} {m}' if m else s


# =============================================================================
# Read growth performance individual data
# =============================================================================

growth = pd.read_csv('growth_performance_tidy.csv')
breeds = ['DLY', 'TFB']
stages = ['15 kg', '45 kg', '75 kg', '105 kg']
adjacent_pairs = [('15 kg', '45 kg'), ('45 kg', '75 kg'), ('75 kg', '105 kg')]

gp_vars = {
    'ADG_kg_d':  'ADG, kg/d',
    'ADFI_kg_d': 'ADFI, kg/d',
    'F_G':       'F/G',
}

# Descriptive stats
gp_stats = {}  # gp_stats[var][breed][stage] = {'n','mean','sd','sem'}
for var in gp_vars:
    gp_stats[var] = {}
    for breed in breeds:
        gp_stats[var][breed] = {}
        for stage in stages:
            d = growth[(growth['Breed'] == breed) & (growth['Stage'] == stage)][var]
            gp_stats[var][breed][stage] = {
                'n': len(d), 'mean': d.mean(), 'sd': d.std(), 'sem': d.sem()
            }

# Breed comparison (Welch's t-test)
gp_breed_p = {}
for var in gp_vars:
    gp_breed_p[var] = {}
    for stage in stages:
        d1 = growth[(growth['Breed'] == 'DLY') & (growth['Stage'] == stage)][var]
        d2 = growth[(growth['Breed'] == 'TFB') & (growth['Stage'] == stage)][var]
        t, p = stats.ttest_ind(d1, d2, equal_var=False)
        pooled_sd = np.sqrt((d1.std()**2 + d2.std()**2) / 2)
        d_cohen = (d1.mean() - d2.mean()) / pooled_sd if pooled_sd > 0 else np.nan
        gp_breed_p[var][stage] = {'t': t, 'p': p, 'cohens_d': d_cohen}

# Adjacent stage comparison within breed (independent t-test; pig IDs unavailable)
# NOTE: original design is longitudinal — paired t-test is more appropriate with pig IDs
gp_adj_p = {}
for var in gp_vars:
    gp_adj_p[var] = {}
    for breed in breeds:
        gp_adj_p[var][breed] = {}
        for s1, s2 in adjacent_pairs:
            d1 = growth[(growth['Breed'] == breed) & (growth['Stage'] == s1)][var]
            d2 = growth[(growth['Breed'] == breed) & (growth['Stage'] == s2)][var]
            t, p = stats.ttest_ind(d1, d2, equal_var=False)
            gp_adj_p[var][breed][f'{s1} vs {s2}'] = {'t': t, 'p': p}


print("Growth performance: done")
for var in gp_vars:
    print(f"  {var}:")
    for stage in stages:
        print(f"    {stage}: DLY={gp_stats[var]['DLY'][stage]['mean']:.3f}±{gp_stats[var]['DLY'][stage]['sem']:.4f}, "
              f"TFB={gp_stats[var]['TFB'][stage]['mean']:.3f}±{gp_stats[var]['TFB'][stage]['sem']:.4f}, "
              f"P={gp_breed_p[var][stage]['p']:.4f}{sig_marker(gp_breed_p[var][stage]['p'])}")

# =============================================================================
# Read source summary Excel (N balance + isotope + digestibility)
# =============================================================================

src = pd.read_excel('phenotype/data nb isotope.xlsx', sheet_name='Sheet2', header=None)

# --- N balance (rows 1-7, 0-indexed) ---
nb_items = [
    'N intake, g/d', 'FN, g/d', 'UN, g/d', 'TN, g/d',
    'RN, g/d', 'N retention, %', 'N ABV, %'
]

# Column mapping: (breed, stage) -> column index
mean_cols = [
    ('DLY','15 kg',1), ('TFB','15 kg',2),
    ('DLY','45 kg',4), ('TFB','45 kg',5),
    ('DLY','75 kg',7), ('TFB','75 kg',8),
    ('DLY','105 kg',10), ('TFB','105 kg',11),
]
breed_p_cols = [('15 kg',3), ('45 kg',6), ('75 kg',9), ('105 kg',12)]
adj_p_cols = [
    ('DLY','15 kg vs 45 kg',14), ('DLY','45 kg vs 75 kg',15), ('DLY','75 kg vs 105 kg',16),
    ('TFB','15 kg vs 45 kg',17), ('TFB','45 kg vs 75 kg',18), ('TFB','75 kg vs 105 kg',19),
]

def parse_section(items, start_row):
    """Parse a data section from source Excel. Returns dict of item -> parsed data."""
    result = {}
    for i, item in enumerate(items):
        row = src.iloc[start_row + i]
        result[item] = {}

        # Mean±SD values
        for breed, stage, col in mean_cols:
            mean, sd = parse_mean_sd(row[col])
            result[item][f'{breed}_{stage}'] = {'mean': mean, 'sd': sd}

        # Breed comparison P-values
        for stage, col in breed_p_cols:
            p_val, _ = parse_pvalue(row[col])
            result[item][f'P_breed_{stage}'] = p_val

        # Adjacent stage P-values
        for breed, pair, col in adj_p_cols:
            p_val, _ = parse_pvalue(row[col])
            result[item][f'P_{breed}_{pair}'] = p_val

    return result

nb_data = parse_section(nb_items, start_row=1)

# --- Isotope tracer (rows 10-14) ---
iso_items = [
    'N flux, g/kg BW^0.75/d',
    'Protein synthesis, N g/kg BW^0.75/d',
    'Protein degradation, N g/kg BW^0.75/d',
    'Protein deposition, N g/kg BW^0.75/d',
    'Protein retention/synthesis, %',
]
iso_data = parse_section(iso_items, start_row=10)

# --- Digestibility (rows 18-20, only 15+45 kg) ---
dig_items = ['DM, %', 'CP, %', 'EE, %']
dig_data = {}
for i, item in enumerate(dig_items):
    row = src.iloc[18 + i]
    dig_data[item] = {}
    for breed, stage, col in [
        ('DLY','15 kg',1), ('TFB','15 kg',2),
        ('DLY','45 kg',4), ('TFB','45 kg',5),
    ]:
        mean, sd = parse_mean_sd(row[col])
        dig_data[item][f'{breed}_{stage}'] = {'mean': mean, 'sd': sd}
    for stage, col in [('15 kg',3), ('45 kg',6)]:
        p_val, _ = parse_pvalue(row[col])
        dig_data[item][f'P_breed_{stage}'] = p_val
    for breed, col in [('DLY',7), ('TFB',8)]:
        p_val, _ = parse_pvalue(row[col])
        dig_data[item][f'P_{breed}_15 kg vs 45 kg'] = p_val


print("\nN balance: done")
print("Isotope: done")
print("Digestibility: done")

# Quick NaN check
for name, data, items in [
    ('N balance', nb_data, nb_items),
    ('Isotope', iso_data, iso_items),
    ('Digestibility', dig_data, dig_items),
]:
    for item in items:
        for key, val in data[item].items():
            if isinstance(val, dict):
                if np.isnan(val['mean']) or np.isnan(val['sd']):
                    print(f"  WARNING: {name} {item} {key}: mean={val['mean']}, sd={val['sd']}")
            elif np.isnan(val):
                print(f"  WARNING: {name} {item} {key}: P={val}")

# =============================================================================
# Build Excel workbook
# =============================================================================

wb = Workbook()

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUBHEADER_FONT = Font(name='Arial', size=10, bold=True)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=12, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_row(ws, row, ncols, bold_first=False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BOLD_FONT if (bold_first and col == 1) else BODY_FONT
        cell.alignment = CENTER if col > 1 else LEFT
        cell.border = THIN_BORDER

def auto_width(ws, min_width=10, max_width=28):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                val = str(cell.value)
                # Rough CJK width
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                lengths.append(width)
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


def write_descriptive_sheet(ws, title, items, item_labels, stats_dict, source_type='individual'):
    """Write a descriptive statistics sheet.

    source_type: 'individual' (mean±SEM from raw) or 'source' (mean±SD from summary file)
    """
    ws.title = title

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1, value=f'{title} — Descriptive Statistics (Mean ± SEM)').font = TITLE_FONT

    # Headers
    headers = ['Item']
    for stage in stages:
        headers.append(f'DLY {stage}')
        headers.append(f'TFB {stage}')
    headers.append('')  # spacer
    for stage in stages:
        headers.append(f'DLY {stage}')
        headers.append(f'TFB {stage}')

    ncols = len(headers)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header_row(ws, 3, ncols)

    # Sub-header for n
    ws.cell(row=4, column=1, value='n')
    col = 2
    for stage in stages:
        for breed in breeds:
            if source_type == 'individual':
                n = stats_dict[items[0]][breed][stage]['n'] if items[0] in stats_dict else ''
                # Actually need per-item n
                pass
            else:
                # Source data: use placeholder, n info from manuscripts
                n_val = '(see manuscript)'
                ws.cell(row=4, column=col, value=n_val)
                ws.cell(row=4, column=col+1, value=n_val)
            col += 2
    style_row(ws, 4, ncols, bold_first=True)

    # Data rows
    for i, item in enumerate(items):
        row = 5 + i
        label = item_labels.get(item, item)
        ws.cell(row=row, column=1, value=label)

        col = 2
        for stage in stages:
            for breed in breeds:
                if source_type == 'individual':
                    d = stats_dict[item][breed][stage]
                    val = format_mean_sem(d['mean'], d['sem'], decimals=3 if 'F/G' in label else 2)
                else:
                    d = stats_dict[item][f'{breed}_{stage}']
                    # Source provides SD; display as mean ± SD
                    val = format_mean_sd(d['mean'], d['sd'], decimals=2)
                ws.cell(row=row, column=col, value=val)
                col += 1
        style_row(ws, row, ncols)

    return ncols  # for positioning P-value tables below


# =============================================================================
# Sheet 1: Growth Performance — Descriptive
# =============================================================================

ws1 = wb.active
ws1.title = '1_Growth_Descriptive'

title = 'Growth Performance'
ws1.merge_cells('A1:J1')
ws1.cell(row=1, column=1, value=f'{title} — Descriptive Statistics').font = TITLE_FONT

ncols = 10
# Row 3: stage headers
ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1)
ws1.cell(row=3, column=1, value='Item')
col = 2
for stage in stages:
    ws1.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws1.cell(row=3, column=col, value=stage)
    col += 2
for ci in range(1, ncols+1):
    cell = ws1.cell(row=3, column=ci)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Row 4: breed sub-headers
ws1.cell(row=4, column=1, value='Breed')
for stage_idx in range(4):
    col = 2 + stage_idx * 2
    for j, breed in enumerate(breeds):
        ws1.cell(row=4, column=col+j, value=breed)
for ci in range(1, ncols+1):
    cell = ws1.cell(row=4, column=ci)
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Data rows
for i, (var, label) in enumerate(gp_vars.items()):
    row = 5 + i
    ws1.cell(row=row, column=1, value=label)
    col = 2
    for stage in stages:
        for breed in breeds:
            d = gp_stats[var][breed][stage]
            val = format_mean_sem(d['mean'], d['sem'], decimals=3)
            ws1.cell(row=row, column=col, value=val)
            col += 1
    style_row(ws1, row, ncols)

# Row 8: n
row_n = 8
ws1.cell(row=row_n, column=1, value='n')
for stage_idx in range(4):
    col = 2 + stage_idx * 2
    for j, breed in enumerate(breeds):
        n = gp_stats['ADG_kg_d'][breed][stages[stage_idx]]['n']
        ws1.cell(row=row_n, column=col+j, value=n)
style_row(ws1, row_n, ncols, bold_first=True)

# Row 9: note
ws1.merge_cells(start_row=9, start_column=1, end_row=9, end_column=ncols)
ws1.cell(row=9, column=1, value='Note: Values are Mean ± SEM. Breed comparison uses Welch\'s independent t-test. '
         'Within-breed adjacent-stage comparison uses independent t-test (pig IDs not in tidy data; paired t-test recommended).').font = Font(name='Arial', size=9, italic=True)

auto_width(ws1)

# =============================================================================
# Sheet 2: Growth Performance — Breed P-values
# =============================================================================

ws2 = wb.create_sheet('2_Growth_Breed_P')

ws2.merge_cells('A1:G1')
ws2.cell(row=1, column=1, value=f'{title} — Breed Comparison (DLY vs TFB) at Each Stage').font = TITLE_FONT

headers2 = ['Item', '15 kg', '45 kg', '75 kg', '105 kg', 'Test', 'Note']
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=ci, value=h)
style_header_row(ws2, 3, len(headers2))

for i, (var, label) in enumerate(gp_vars.items()):
    row = 4 + i
    ws2.cell(row=row, column=1, value=label)
    for j, stage in enumerate(stages):
        p = gp_breed_p[var][stage]
        val = f'P = {format_p(p["p"])}'
        if not np.isnan(p['cohens_d']):
            val += f'\nd = {p["cohens_d"]:.3f}'
        ws2.cell(row=row, column=2+j, value=val)
    ws2.cell(row=row, column=6, value="Welch's t")
    ws2.cell(row=row, column=7, value='')
    style_row(ws2, row, len(headers2))

auto_width(ws2)

# =============================================================================
# Sheet 3: Growth Performance — Adjacent Stage P-values
# =============================================================================

ws3 = wb.create_sheet('3_Growth_Adjacent_P')

ws3.merge_cells('A1:H1')
ws3.cell(row=1, column=1, value=f'{title} — Within-Breed Adjacent Stage Comparison').font = TITLE_FONT

src_label_map = {
    'ADG_kg_d': 'ADG, kg/d',
    'ADFI_kg_d': 'ADFI, kg/d',
    'F_G': 'F/G'
}

headers3 = ['Item', 'Breed', '15 vs 45 kg', '45 vs 75 kg', '75 vs 105 kg', 'Test', 'Note']
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=ci, value=h)
style_header_row(ws3, 3, len(headers3))

row = 4
for var, label in src_label_map.items():
    for breed in breeds:
        ws3.cell(row=row, column=1, value=label)
        ws3.cell(row=row, column=2, value=breed)
        for j, (s1, s2) in enumerate(adjacent_pairs):
            p = gp_adj_p[var][breed][f'{s1} vs {s2}']['p']
            ws3.cell(row=row, column=3+j, value=f'P = {format_p(p)}')
        ws3.cell(row=row, column=6, value='Independent t')
        ws3.cell(row=row, column=7, value='Same-breed pigs, different stages')
        style_row(ws3, row, len(headers3))
        row += 1

auto_width(ws3)


# =============================================================================
# Generic sheet builder for N balance / Isotope (3 sheets each)
# =============================================================================

def write_triple_sheets(wb, title_short, items, item_labels, data_dict, base_name,
                        source_type='source', decimals=2):
    """Write 3 sheets: Descriptive, Breed P, Adjacent Stage P.

    source_type: 'source' = mean±SD from summary file, display as mean±SD
    """
    # --- Descriptive ---
    ws = wb.create_sheet(f'{base_name}_Descriptive')
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value=f'{title_short} — Descriptive Statistics').font = TITLE_FONT

    ncols = 9
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1)
    ws.cell(row=3, column=1, value='Item')
    col = 2
    for stage in stages:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(row=3, column=col, value=stage)
        col += 2
    style_header_row(ws, 3, ncols)

    # Row 4: breed sub-headers
    ws.cell(row=4, column=1, value='Breed')
    for stage_idx in range(4):
        col = 2 + stage_idx * 2
        for j, breed in enumerate(breeds):
            ws.cell(row=4, column=col+j, value=breed)
    for ci in range(1, ncols+1):
        cell = ws.cell(row=4, column=ci)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for i, item in enumerate(items):
        row = 5 + i
        label = item_labels.get(item, item)
        ws.cell(row=row, column=1, value=label)
        col = 2
        for stage in stages:
            for breed in breeds:
                d = data_dict[item][f'{breed}_{stage}']
                val = format_mean_sd(d['mean'], d['sd'], decimals=decimals)
                ws.cell(row=row, column=col, value=val)
                col += 1
        style_row(ws, row, ncols)

    # Note
    note_row = 5 + len(items) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
    ws.cell(row=note_row, column=1,
            value='Note: Values are Mean ± SD (from summary source). '
                  'Individual data needed for Mean ± SEM and independent P-value recalculation.').font = Font(name='Arial', size=9, italic=True)
    auto_width(ws)

    # --- Breed P ---
    ws = wb.create_sheet(f'{base_name}_Breed_P')
    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value=f'{title_short} — Breed Comparison (DLY vs TFB)').font = TITLE_FONT

    headers = ['Item', '15 kg', '45 kg', '75 kg', '105 kg', 'Test', 'Note']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header_row(ws, 3, len(headers))

    for i, item in enumerate(items):
        row = 4 + i
        label = item_labels.get(item, item)
        ws.cell(row=row, column=1, value=label)
        for j, stage in enumerate(stages):
            p = data_dict[item][f'P_breed_{stage}']
            ws.cell(row=row, column=2+j, value=format_p(p))
        ws.cell(row=row, column=6, value='See source')
        ws.cell(row=row, column=7, value='')
        style_row(ws, row, len(headers))
    auto_width(ws)

    # --- Adjacent Stage P ---
    ws = wb.create_sheet(f'{base_name}_Adjacent_P')
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value=f'{title_short} — Within-Breed Adjacent Stage Comparison').font = TITLE_FONT

    headers = ['Item', 'Breed', '15 vs 45 kg', '45 vs 75 kg', '75 vs 105 kg', 'Test', 'Note']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for item in items:
        label = item_labels.get(item, item)
        for breed in breeds:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=breed)
            for j, (s1, s2) in enumerate(adjacent_pairs):
                key = f'P_{breed}_{s1} vs {s2}'
                p_val = data_dict[item].get(key, np.nan)
                ws.cell(row=row, column=3+j, value=format_p(p_val))
            ws.cell(row=row, column=6, value='See source')
            ws.cell(row=row, column=7, value='From original summary data')
            style_row(ws, row, len(headers))
            row += 1
    auto_width(ws)


# N balance label map
nb_labels = {
    'N intake, g/d': 'N intake, g/d',
    'FN, g/d': 'Fecal N, g/d',
    'UN, g/d': 'Urinary N, g/d',
    'TN, g/d': 'Total N excretion, g/d',
    'RN, g/d': 'Retained N, g/d',
    'N retention, %': 'N retention rate, %',
    'N ABV, %': 'N apparent biological value, %',
}

write_triple_sheets(wb, 'N Balance', nb_items, nb_labels, nb_data, '4_NBalance')

# Isotope label map
iso_labels = {
    'N flux, g/kg BW^0.75/d': 'N flux, g/kg BW^0.75/d',
    'Protein synthesis, N g/kg BW^0.75/d': 'Protein synthesis, g/kg BW^0.75/d',
    'Protein degradation, N g/kg BW^0.75/d': 'Protein degradation, g/kg BW^0.75/d',
    'Protein deposition, N g/kg BW^0.75/d': 'Protein deposition, g/kg BW^0.75/d',
    'Protein retention/synthesis, %': 'Retention / Synthesis, %',
}
write_triple_sheets(wb, 'Isotope Tracer', iso_items, iso_labels, iso_data, '5_Isotope', decimals=2)


# =============================================================================
# Digestibility sheet (only 15+45 kg)
# =============================================================================

dig_labels = {'DM, %': 'DM digestibility, %', 'CP, %': 'CP digestibility, %', 'EE, %': 'EE digestibility, %'}

ws = wb.create_sheet('6_Digestibility')
ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value='Nutrient Digestibility — 15 & 45 kg').font = TITLE_FONT

# Descriptive section
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1)
ws.cell(row=3, column=1, value='Item')
ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
ws.cell(row=3, column=2, value='15 kg')
ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
ws.cell(row=3, column=4, value='45 kg')
ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=6)
ws.cell(row=3, column=6, value='')

ncols_d = 8
style_header_row(ws, 3, ncols_d)

ws.cell(row=4, column=1, value='Breed')
for col_base in [2, 4]:
    for j, breed in enumerate(breeds):
        ws.cell(row=4, column=col_base + j, value=breed)
ws.cell(row=4, column=6, value='P Breed 15kg')
ws.cell(row=4, column=7, value='P Breed 45kg')
ws.cell(row=4, column=8, value='P DLY 15vs45')
for ci in range(1, ncols_d+1):
    cell = ws.cell(row=4, column=ci)
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, item in enumerate(dig_items):
    row = 5 + i
    label = dig_labels.get(item, item)
    ws.cell(row=row, column=1, value=label)
    for stage_idx, stage in enumerate(['15 kg', '45 kg']):
        col = 2 + stage_idx * 2
        for j, breed in enumerate(breeds):
            d = dig_data[item][f'{breed}_{stage}']
            ws.cell(row=row, column=col+j, value=format_mean_sd(d['mean'], d['sd']))
    ws.cell(row=row, column=6, value=format_p(dig_data[item]['P_breed_15 kg']))
    ws.cell(row=row, column=7, value=format_p(dig_data[item]['P_breed_45 kg']))
    ws.cell(row=row, column=8, value=format_p(dig_data[item].get('P_DLY_15 kg vs 45 kg', np.nan)))
    style_row(ws, row, ncols_d)

# TFB adjacent P
ws.cell(row=8, column=1, value='')
ws.cell(row=9, column=8, value='P TFB 15vs45')
ws.cell(row=9, column=8).font = BOLD_FONT
ws.cell(row=9, column=8).border = THIN_BORDER
for i, item in enumerate(dig_items):
    row = 10 + i
    ws.cell(row=row, column=8, value=format_p(dig_data[item].get('P_TFB_15 kg vs 45 kg', np.nan)))
    ws.cell(row=row, column=8).border = THIN_BORDER
    ws.cell(row=row, column=8).alignment = CENTER

auto_width(ws)


# =============================================================================
# Summary: P-value matrix (all in one place)
# =============================================================================

ws = wb.create_sheet('7_Pvalue_Summary')

ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value='Complete P-value Matrix — Breed Comparison & Adjacent Stage Comparison').font = TITLE_FONT

r = 3
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='Breed Comparison: DLY vs TFB at each stage').font = BOLD_FONT
r += 1

pv_headers = ['Item', 'Type', '15 kg', '45 kg', '75 kg', '105 kg']
for ci, h in enumerate(pv_headers, 1):
    ws.cell(row=r, column=ci, value=h)
style_header_row(ws, r, len(pv_headers))
r += 1

for label, items, data_dict, dtype in [
    ('Growth Performance', list(gp_vars.keys()), None, 'growth'),
    ('N Balance', nb_items, nb_data, 'source'),
    ('Isotope Tracer', iso_items, iso_data, 'source'),
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(pv_headers))
    ws.cell(row=r, column=1, value=label)
    for ci in range(1, len(pv_headers)+1):
        cell = ws.cell(row=r, column=ci)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
    r += 1

    for item in items:
        if dtype == 'growth':
            item_label = gp_vars[item]
            ws.cell(row=r, column=1, value=item_label)
            ws.cell(row=r, column=2, value='Breed P')
            for j, stage in enumerate(stages):
                p = gp_breed_p[item][stage]['p']
                ws.cell(row=r, column=3+j, value=format_p(p))
            style_row(ws, r, len(pv_headers))
            r += 1
        else:
            item_label = item
            ws.cell(row=r, column=1, value=item_label)
            ws.cell(row=r, column=2, value='Breed P')
            for j, stage in enumerate(stages):
                p = data_dict[item][f'P_breed_{stage}']
                ws.cell(row=r, column=3+j, value=format_p(p))
            style_row(ws, r, len(pv_headers))
            r += 1

# Adjacent stage section
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(row=r, column=1, value='Adjacent Stage Comparison (within breed)').font = BOLD_FONT
r += 1

adj_headers = ['Item', 'Breed', '15 vs 45 kg', '45 vs 75 kg', '75 vs 105 kg']
for ci, h in enumerate(adj_headers, 1):
    ws.cell(row=r, column=ci, value=h)
style_header_row(ws, r, len(adj_headers))
r += 1

for section_label, items, data_dict, dtype in [
    ('Growth Performance', list(gp_vars.keys()), None, 'growth'),
    ('N Balance', nb_items, nb_data, 'source'),
    ('Isotope Tracer', iso_items, iso_data, 'source'),
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(adj_headers))
    ws.cell(row=r, column=1, value=section_label)
    for ci in range(1, len(adj_headers)+1):
        cell = ws.cell(row=r, column=ci)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
    r += 1

    for item in items:
        if dtype == 'growth':
            item_label = gp_vars[item]
            for breed in breeds:
                ws.cell(row=r, column=1, value=item_label)
                ws.cell(row=r, column=2, value=breed)
                for j, (s1, s2) in enumerate(adjacent_pairs):
                    p = gp_adj_p[item][breed][f'{s1} vs {s2}']['p']
                    ws.cell(row=r, column=3+j, value=format_p(p))
                style_row(ws, r, len(adj_headers))
                r += 1
        else:
            item_label = item
            for breed in breeds:
                ws.cell(row=r, column=1, value=item_label)
                ws.cell(row=r, column=2, value=breed)
                for j, (s1, s2) in enumerate(adjacent_pairs):
                    key = f'P_{breed}_{s1} vs {s2}'
                    p_val = data_dict[item].get(key, np.nan)
                    ws.cell(row=r, column=3+j, value=format_p(p_val))
                style_row(ws, r, len(adj_headers))
                r += 1

auto_width(ws)


# =============================================================================
# Save
# =============================================================================

output_path = 'integrated_15_105kg_all_phenotypes.xlsx'
wb.save(output_path)
print(f'\n{"="*60}')
print(f'Successfully saved: {output_path}')
print(f'Sheets: {wb.sheetnames}')
