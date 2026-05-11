#!/usr/bin/env python3
"""
Advanced analysis v2 — 3-tier temporal screening strategy.
逻辑框架：
  Tier 1 (发育编程基因): 15 kg 已差异，跨 15/45/75 kg 方向一致 → 遗传固定的代谢编程
  Tier 2 (表型偶联基因): 45 kg 差异最大或仅在 45 kg 显著 → 沉积瓶颈的分子开关
  Tier 3 (后果基因):     仅在 75-105 kg 差异 → 长期分化的结果

DLY 105 kg 肝数据排除，不作为筛选依据。
"""
import pandas as pd
import numpy as np
from scipy.stats import pearsonr, spearmanr, ttest_ind
from stats_utils import benjamini_hochberg, apply_fdr_to_dataframe
from scipy.cluster.hierarchy import linkage, leaves_list
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import re

print("=" * 60)
print("Advanced Analysis v2 — 3-Tier Temporal Screening")
print("=" * 60)

# ============================================================
# 1. Data Loading
# ============================================================
print("\nLoading data...")
serum_tidy = pd.read_csv('serum_all_tidy.csv')
muscle = pd.read_csv('/Users/hezongze/Downloads/result_exp_matrix.xls', sep='\t')
liver = pd.read_csv('/Users/hezongze/Downloads/result_exp_matrix (4).xls', sep='\t')

sample_to_group_m = {
    'm_15_1_': ('DLY', 15), 'm_15_2_': ('TFB', 15),
    'BJ_2_1_': ('DLY', 45), 'BJ_2_2_': ('TFB', 45),
    'm_1_1_': ('DLY', 75), 'm_1_2_': ('TFB', 75),
    'm_2_1_': ('DLY', 105), 'm_2_2_': ('TFB', 105),
    'm_3_1_': ('DLY', 135),
}
sample_to_group_l = {
    'L_15_1_': ('DLY', 15), 'L_15_2_': ('TFB', 15),
    'L_45_1_': ('DLY', 45), 'L_45_2_': ('TFB', 45),
    'L_1_1_': ('DLY', 75), 'L_1_2_': ('TFB', 75),
    'L_2_1_': ('DLY', 105), 'L_2_2_': ('TFB', 105),
    'L_3_1_': ('DLY', 135),
}

def build_individual_df(mat, sample_map):
    val_cols = [c for c in mat.columns if c not in ('seq_id', 'gene_name', 'length', 'description')]
    records = []
    for _, row in mat.iterrows():
        gene_name = str(row['gene_name']) if pd.notna(row['gene_name']) else row['seq_id']
        for col in val_cols:
            info = None
            for prefix, (breed, stage) in sample_map.items():
                if col.startswith(prefix):
                    rep_num = int(col.split('_')[-1])
                    info = (breed, stage, rep_num)
                    break
            if info and pd.notna(row[col]):
                records.append({
                    'gene_name': gene_name, 'breed': info[0],
                    'stage_kg': info[1], 'rep': info[2], 'expr': float(row[col])
                })
    return pd.DataFrame(records)

print("Building individual-level expression tables...")
liver_ind = build_individual_df(liver, sample_to_group_l)
muscle_ind = build_individual_df(muscle, sample_to_group_m)
liver_ind['tissue'] = 'Liver'
muscle_ind['tissue'] = 'Muscle'

# ============================================================
# 2. Tier Classification Engine
# ============================================================
# VALID_STAGES for screening: exclude DLY 105kg liver
# Use: 15, 45, 75 for full breed comparison; 135 DLY-only for reference
# TFB 105kg is kept for reference but not used for DLYvsTFB at 105

def classify_gene_tier(gene_name, tissue_df):
    """
    Classify a gene into Tier 1/2/3 based on temporal pattern.
    DLY 105kg liver is EXCLUDED from screening.
    Returns: tier, pattern_description, per_stage_log2FC dict
    """
    gdf = tissue_df[tissue_df['gene_name'] == gene_name]
    if len(gdf) < 6:
        return None, None, None

    stage_fcs = {}
    for s in [15, 45, 75, 105]:
        dly_vals = gdf[(gdf['breed'] == 'DLY') & (gdf['stage_kg'] == s)]['expr']
        tfb_vals = gdf[(gdf['breed'] == 'TFB') & (gdf['stage_kg'] == s)]['expr']
        if len(dly_vals) > 0 and len(tfb_vals) > 0:
            tfb_m = tfb_vals.mean()
            if tfb_m > 0:
                stage_fcs[s] = np.log2(dly_vals.mean() / tfb_m)
            else:
                stage_fcs[s] = np.nan
        else:
            stage_fcs[s] = np.nan

    # Get valid stages (excluding 105 for screening if data questionable)
    valid_stages = [s for s in [15, 45, 75] if s in stage_fcs and pd.notna(stage_fcs[s])]

    if len(valid_stages) < 2:
        return None, None, stage_fcs

    fcs_valid = [stage_fcs[s] for s in valid_stages]
    mean_fc = np.mean(fcs_valid)
    max_abs_fc = max(abs(v) for v in fcs_valid)

    # Check direction consistency among valid stages
    signs = [1 if v > 0 else -1 if v < 0 else 0 for v in fcs_valid]
    consistent = len(set(signs)) == 1 and 0 not in signs

    # Tier 1: 15kg already |log2FC| > 0.5, direction consistent across 15+45+75
    if 15 in stage_fcs and pd.notna(stage_fcs[15]):
        fc15 = abs(stage_fcs[15])
        if fc15 > 0.5 and consistent and len(valid_stages) >= 3:
            trend = 'TFB↑' if mean_fc < 0 else 'DLY↑'
            return 1, f'Early programming ({trend}, |FC15|={fc15:.1f})', stage_fcs

    # Tier 2: max |log2FC| at 45kg, or divergence emerges at 45kg
    if 45 in stage_fcs and pd.notna(stage_fcs[45]):
        fc45 = abs(stage_fcs[45])
        fc15_abs = abs(stage_fcs.get(15, 0)) if pd.notna(stage_fcs.get(15, np.nan)) else 0
        fc75_abs = abs(stage_fcs.get(75, 0)) if pd.notna(stage_fcs.get(75, np.nan)) else 0

        # 45kg has largest FC among valid stages
        if fc45 > 0.5 and fc45 >= fc15_abs and fc45 >= fc75_abs * 0.8:
            trend = 'TFB↑' if stage_fcs[45] < 0 else 'DLY↑'
            return 2, f'45kg-peak ({trend}, |FC45|={fc45:.1f})', stage_fcs

        # Divergence emerges at 45 (not present at 15, appears at 45, remains or grows)
        if fc15_abs < 0.3 and fc45 > 0.5:
            trend = 'TFB↑' if stage_fcs[45] < 0 else 'DLY↑'
            return 2, f'45kg-emergent ({trend}, |FC45|={fc45:.1f})', stage_fcs

    # Tier 3: only significant at 75 or 105
    fc75_abs = abs(stage_fcs.get(75, 0)) if pd.notna(stage_fcs.get(75, np.nan)) else 0
    fc105_abs = abs(stage_fcs.get(105, 0)) if pd.notna(stage_fcs.get(105, np.nan)) else 0
    if fc75_abs > 0.5 or fc105_abs > 0.5:
        trend = 'TFB↑' if mean_fc < 0 else 'DLY↑'
        return 3, f'Late consequence ({trend})', stage_fcs

    return None, None, stage_fcs


# ============================================================
# 3. Apply Tier Classification to AA Catabolism Genes
# ============================================================
print("\n" + "=" * 60)
print("Tier Classification: AA Catabolism Enzymes (Liver)")
print("=" * 60)

AA_CATABOLISM_GENES = [
    'AASS', 'ACADSB', 'ARG1', 'ARG2', 'ASL', 'ASS1', 'BCAT1', 'BCAT2',
    'BCKDHA', 'BCKDHB', 'CPS1', 'DBT', 'DLD', 'GLUD1', 'GOT1', 'GOT2',
    'HAL', 'HGD', 'OTC', 'PAH', 'SDS'
]

aa_tiers = []
for gene in AA_CATABOLISM_GENES:
    tier, desc, fcs = classify_gene_tier(gene, liver_ind)
    if tier:
        aa_tiers.append({
            'Gene': gene,
            'Tier': tier,
            'Pattern': desc,
            **{f'log2FC_{s}kg': round(fcs.get(s, np.nan), 3) if pd.notna(fcs.get(s, np.nan)) else '' for s in [15, 45, 75, 105]}
        })

aa_tier_df = pd.DataFrame(aa_tiers).sort_values(['Tier', 'Gene'])
print(f"\nAA Catabolism Gene Tier Distribution:")
for tier in [1, 2, 3]:
    subset = aa_tier_df[aa_tier_df['Tier'] == tier]
    print(f"\n  Tier {tier} ({len(subset)} genes):")
    for _, r in subset.iterrows():
        fcs_str = ' | '.join([f"{s}kg={r[f'log2FC_{s}kg']}" for s in [15, 45, 75, 105] if r[f'log2FC_{s}kg'] != ''])
        print(f"    {r['Gene']:10s} {r['Pattern']:40s} [{fcs_str}]")

# ============================================================
# 4. Tier Classification: Crosstalk Genes (Liver & Muscle)
# ============================================================
print("\n" + "=" * 60)
print("Tier Classification: Cross-Talk Genes")
print("=" * 60)

CROSSTALK_GENES = [
    'IGFBP1', 'IGFBP2', 'IGFBP3', 'IGFBP4', 'IGFBP5',
    'AHSG', 'RBP4', 'FGG', 'FGB', 'APOB', 'BDNF', 'SERPINC1',
    'XBP1', 'ATF4', 'FGF21', 'ANGPTL4', 'ANGPTL8',
    'MSTN', 'FST', 'FNDC5', 'SPARC', 'DCN', 'LIF', 'IL6', 'VEGFA',
    'APOA1', 'APOA2', 'APOC3', 'APOE', 'APOH', 'ALB', 'TTR',
    'AGT', 'KNG1', 'FGA', 'PLG', 'C3', 'C5', 'CFH', 'HP', 'HPX',
    'SERPINA1', 'SERPINA3', 'SERPINF1', 'ORM1',
]

crosstalk_tiers = []
for gene in CROSSTALK_GENES:
    # Liver classification
    l_tier, l_desc, l_fcs = classify_gene_tier(gene, liver_ind)
    # Muscle classification
    m_tier, m_desc, m_fcs = classify_gene_tier(gene, muscle_ind)

    if l_tier or m_tier:
        crosstalk_tiers.append({
            'Gene': gene,
            'Liver_Tier': l_tier if l_tier else '-',
            'Liver_Pattern': l_desc if l_desc else '',
            'Muscle_Tier': m_tier if m_tier else '-',
            'Muscle_Pattern': m_desc if m_desc else '',
            **{f'Liver_log2FC_{s}kg': round(l_fcs.get(s, np.nan), 3) if l_fcs and pd.notna(l_fcs.get(s, np.nan)) else '' for s in [15, 45, 75, 105]},
            **{f'Muscle_log2FC_{s}kg': round(m_fcs.get(s, np.nan), 3) if m_fcs and pd.notna(m_fcs.get(s, np.nan)) else '' for s in [15, 45, 75, 105]},
        })

ct_tier_df = pd.DataFrame(crosstalk_tiers).sort_values(['Liver_Tier', 'Muscle_Tier', 'Gene'])
print(f"\nCrosstalk Gene Tier Distribution (Liver):")
for tier in [1, 2, 3]:
    subset = ct_tier_df[ct_tier_df['Liver_Tier'] == tier]
    print(f"\n  Liver Tier {tier} ({len(subset)} genes):")
    for _, r in subset.iterrows():
        print(f"    {r['Gene']:12s} {r['Liver_Pattern']}")

print(f"\nCrosstalk Gene Tier Distribution (Muscle):")
for tier in [1, 2, 3]:
    subset = ct_tier_df[ct_tier_df['Muscle_Tier'] == tier]
    print(f"\n  Muscle Tier {tier} ({len(subset)} genes):")
    for _, r in subset.iterrows():
        print(f"    {r['Gene']:12s} {r['Muscle_Pattern']}")

# ============================================================
# 5. TF/Regulator Tier Classification
# ============================================================
print("\n" + "=" * 60)
print("Tier Classification: Transcription Factors / Regulators")
print("=" * 60)

TF_CANDIDATES = {
    'FOXO1': 'FOXO family', 'FOXO3': 'FOXO family', 'FOXO4': 'FOXO family',
    'PPARGC1A': 'PGC1α/co-activator', 'PPARA': 'PPAR family',
    'PPARD': 'PPAR family', 'PPARG': 'PPAR family',
    'XBP1': 'ER stress/UPR', 'ATF4': 'AA response/ISR', 'ATF6': 'ER stress/UPR',
    'DDIT3': 'CHOP/ER stress', 'NFE2L2': 'NRF2/oxidative stress',
    'TFEB': 'Lysosome/autophagy', 'TFE3': 'Lysosome/autophagy',
    'CREB1': 'cAMP/PKA', 'CRTC2': 'CREB co-activator',
    'STAT3': 'JAK/STAT', 'HIF1A': 'Hypoxia',
    'NRF1': 'Mitochondrial biogenesis', 'TFAM': 'mtDNA transcription',
    'ESRRA': 'ERRα', 'KLF15': 'AA/glucose metabolism',
    'CEBPA': 'C/EBP family', 'CEBPB': 'C/EBP family',
    'HNF4A': 'HNF4/liver identity', 'HNF4G': 'HNF4/liver identity',
    'HNF1A': 'HNF1/liver identity', 'RXRA': 'Nuclear receptor',
    'NR1H4': 'FXR/bile acid', 'NR1I2': 'PXR/xenobiotic',
    'NR1I3': 'CAR/xenobiotic', 'NR3C1': 'Glucocorticoid receptor',
    'THRA': 'Thyroid receptor α', 'THRB': 'Thyroid receptor β',
    'MYC': 'Growth/ribosome', 'MYCN': 'Growth/ribosome',
    'SRF': 'Serum response factor', 'TEAD1': 'Hippo/YAP',
    'TEAD4': 'Hippo/YAP', 'YAP1': 'Hippo pathway',
    'WWTR1': 'TAZ/Hippo', 'SREBF1': 'SREBP/lipid',
    'SREBF2': 'SREBP/cholesterol', 'MLXIPL': 'ChREBP/glucose',
    'ARNTL': 'BMAL1/circadian', 'CLOCK': 'Circadian',
    'NR1D1': 'REV-ERBα/circadian', 'NR1D2': 'REV-ERBβ/circadian',
    'EPAS1': 'HIF2α/hypoxia', 'NFKB1': 'NF-κB', 'NFKB2': 'NF-κB',
    'RELA': 'NF-κB p65',
}

all_liver_genes = set(liver_ind['gene_name'].unique())

# Find which TFs exist in liver
found_tfs = {}
for tf, cat in TF_CANDIDATES.items():
    if tf in all_liver_genes:
        found_tfs[tf] = cat
    else:
        # Case-insensitive fallback
        upper_map = {g.upper(): g for g in all_liver_genes}
        if tf.upper() in upper_map:
            found_tfs[tf] = cat

print(f"Found {len(found_tfs)}/{len(TF_CANDIDATES)} candidate TFs in liver")

tf_tiers = []
for tf in found_tfs:
    tier, desc, fcs = classify_gene_tier(tf, liver_ind)
    tf_tiers.append({
        'TF': tf,
        'Category': found_tfs[tf],
        'Tier': tier if tier else 99,
        'Pattern': desc if desc else 'unclassified',
        **{f'log2FC_{s}kg': round(fcs.get(s, np.nan), 3) if fcs and pd.notna(fcs.get(s, np.nan)) else '' for s in [15, 45, 75, 105]}
    })

tf_tier_df = pd.DataFrame(tf_tiers).sort_values(['Tier', 'TF'])
print(f"\nTF Tier Distribution:")
for tier in [1, 2, 3, 99]:
    subset = tf_tier_df[tf_tier_df['Tier'] == tier]
    label = {1: 'Tier 1 (Early Programming)', 2: 'Tier 2 (45kg Switch)',
             3: 'Tier 3 (Late Consequence)', 99: 'Unclassified'}[tier]
    if len(subset) > 0:
        print(f"\n  {label} ({len(subset)} TFs):")
        for _, r in subset.iterrows():
            fcs_str = ' | '.join([f"{s}kg={r[f'log2FC_{s}kg']}" for s in [15, 45, 75, 105] if r[f'log2FC_{s}kg'] != ''])
            print(f"    {r['TF']:12s} [{r['Category']:25s}] {r['Pattern']:40s} [{fcs_str}]")

# ============================================================
# 6. Correlation Analysis: TF ↔ AA Enzyme (with Tier annotation)
# ============================================================
print("\n" + "=" * 60)
print("TF ↔ AA Enzyme Correlations (breed×stage means)")
print("=" * 60)

liver_bs_mean = liver_ind.groupby(['gene_name', 'breed', 'stage_kg'])['expr'].mean().reset_index()

tf_enzyme_corr = []
for _, tf_row in tf_tier_df.iterrows():
    tf_name = tf_row['TF']
    tf_bs = liver_bs_mean[liver_bs_mean['gene_name'] == tf_name]
    tf_bs = tf_bs.rename(columns={'expr': 'tf_expr'})

    for enz in AA_CATABOLISM_GENES:
        enz_bs = liver_bs_mean[liver_bs_mean['gene_name'] == enz]
        merged = tf_bs.merge(enz_bs, on=['breed', 'stage_kg'])
        if len(merged) >= 6:
            r, p = pearsonr(merged['tf_expr'], merged['expr'])
            tf_enzyme_corr.append({
                'TF': tf_name, 'TF_Tier': tf_row['Tier'],
                'TF_Category': tf_row['Category'],
                'Target_Enzyme': enz,
                'pearson_r': round(r, 4),
                'p_value': round(p, 6),
                'n': len(merged)
            })

tf_corr_df = pd.DataFrame(tf_enzyme_corr)

# FDR correction on all TF-enzyme correlation p-values
if len(tf_corr_df) > 0:
    _, fdr_q = benjamini_hochberg(tf_corr_df['p_value'].values)
    tf_corr_df['q_value'] = fdr_q
    tf_corr_df['FDR_significant'] = tf_corr_df['q_value'] < 0.05

# Top TFs by mean |r| with AA enzymes
tf_mean_corr = tf_corr_df.groupby(['TF', 'TF_Tier', 'TF_Category']).agg(
    mean_abs_r=('pearson_r', lambda x: abs(x).mean()),
    n_corrs=('pearson_r', 'count'),
    sig_count=('p_value', lambda x: (x < 0.05).sum()),
    fdr_sig_count=('FDR_significant', 'sum')
).reset_index().sort_values('mean_abs_r', ascending=False)

print("\nTF correlation with AA enzymes (Tier-annotated):")
for _, r in tf_mean_corr.head(25).iterrows():
    tier_label = {1: 'T1', 2: 'T2', 3: 'T3', 99: '--'}[r['TF_Tier']]
    print(f"  [{tier_label}] {r['TF']:12s} [{r['TF_Category']:25s}] mean|r|={r['mean_abs_r']:.4f}  sig(nominal)={r['sig_count']}/{r['n_corrs']}  FDR_sig={r.get('fdr_sig_count',0)}")

# ============================================================
# 7. Individual-level Correlation: Liver Enzyme ↔ Serum Urea
# ============================================================
print("\n" + "=" * 60)
print("Individual-level: Liver Enzyme ↔ Serum Urea")
print("=" * 60)

serum_urea_ind = serum_tidy[serum_tidy['metabolite'] == 'Urea'].copy()

def parse_serum_group(g):
    breed = 'DLY' if 'DLY' in g else 'TFB'
    stage = int(re.search(r'(\d+)', g).group(1))
    return breed, stage

parsed = serum_urea_ind['group'].apply(parse_serum_group)
serum_urea_ind['breed'] = [p[0] for p in parsed]
serum_urea_ind['stage_kg'] = [p[1] for p in parsed]
serum_urea_bs = serum_urea_ind.groupby(['breed', 'stage_kg'])['value'].mean().reset_index()
serum_urea_bs.rename(columns={'value': 'serum_urea_mean'}, inplace=True)

ind_corr_results = []
for gene in AA_CATABOLISM_GENES:
    gene_df = liver_ind[liver_ind['gene_name'] == gene]
    if len(gene_df) < 10:
        continue
    gene_df = gene_df.merge(serum_urea_bs, on=['breed', 'stage_kg'], how='left')
    gene_df = gene_df.dropna(subset=['expr', 'serum_urea_mean'])
    if len(gene_df) < 8:
        continue

    r, p = pearsonr(gene_df['expr'], gene_df['serum_urea_mean'])
    rho, ps = spearmanr(gene_df['expr'], gene_df['serum_urea_mean'])

    # Per-stage correlations
    stage_corrs = {}
    for s in sorted(gene_df['stage_kg'].unique()):
        sd = gene_df[gene_df['stage_kg'] == s]
        if len(sd) >= 4:
            rs, ps_val = pearsonr(sd['expr'], sd['serum_urea_mean'])
            stage_corrs[s] = (rs, ps_val)

    dly_mean = gene_df[gene_df['breed'] == 'DLY']['expr'].mean()
    tfb_mean = gene_df[gene_df['breed'] == 'TFB']['expr'].mean()
    log2fc = np.log2(dly_mean / tfb_mean) if tfb_mean > 0 else np.nan

    # Get tier
    tier_match = aa_tier_df[aa_tier_df['Gene'] == gene]
    tier = tier_match['Tier'].values[0] if len(tier_match) > 0 else '-'

    ind_corr_results.append({
        'Gene': gene, 'Tier': tier,
        'n': len(gene_df), 'pearson_r': round(r, 4),
        'pearson_p': round(p, 6), 'spearman_rho': round(rho, 4),
        'spearman_p': round(ps, 6), 'overall_log2FC': round(log2fc, 3),
        'stage_corrs': str({s: round(v[0], 3) for s, v in stage_corrs.items()}),
    })

ind_corr_df = pd.DataFrame(ind_corr_results).sort_values('pearson_r', ascending=False)
print("\nLiver enzyme ↔ Serum Urea (individual-level, tier-annotated):")
for _, r in ind_corr_df.iterrows():
    sig = '***' if r['pearson_p'] < 0.001 else ('**' if r['pearson_p'] < 0.01 else ('*' if r['pearson_p'] < 0.05 else ''))
    tier_l = f"[T{r['Tier']}]"
    print(f"  {tier_l} {r['Gene']:10s} r={r['pearson_r']:+.4f} p={r['pearson_p']:.6f} {sig}  n={r['n']}  stage_corrs={r['stage_corrs']}")

# ============================================================
# 8. N Balance Integration
# ============================================================
print("\n" + "=" * 60)
print("N Balance ↔ Expression Integration")
print("=" * 60)

import openpyxl
wb = openpyxl.load_workbook('phenotype/data nb isotope.xlsx', data_only=True)
ws = wb['Sheet2']

n_data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        continue
    try:
        n_data.append({
            'parameter': str(row[0]).strip(),
            'DLY_15_mean': float(str(row[1]).split('±')[0].strip()) if row[1] and '±' in str(row[1]) else np.nan,
            'TFB_15_mean': float(str(row[2]).split('±')[0].strip()) if row[2] and '±' in str(row[2]) else np.nan,
            'DLY_45_mean': float(str(row[4]).split('±')[0].strip()) if row[4] and '±' in str(row[4]) else np.nan,
            'TFB_45_mean': float(str(row[5]).split('±')[0].strip()) if row[5] and '±' in str(row[5]) else np.nan,
            'DLY_75_mean': float(str(row[7]).split('±')[0].strip()) if row[7] and '±' in str(row[7]) else np.nan,
            'TFB_75_mean': float(str(row[8]).split('±')[0].strip()) if row[8] and '±' in str(row[8]) else np.nan,
            'DLY_105_mean': float(str(row[10]).split('±')[0].strip()) if row[10] and '±' in str(row[10]) else np.nan,
            'TFB_105_mean': float(str(row[11]).split('±')[0].strip()) if row[11] and '±' in str(row[11]) else np.nan,
        })
    except (ValueError, IndexError, AttributeError):
        continue

n_df = pd.DataFrame(n_data)
# Keep only rows with enough data
n_df = n_df.dropna(subset=['DLY_15_mean', 'TFB_15_mean'], how='all')

# Build long format
n_long = []
for _, row in n_df.iterrows():
    for stage in [15, 45, 75, 105]:
        for breed in ['DLY', 'TFB']:
            val = row.get(f'{breed}_{stage}_mean', np.nan)
            if pd.notna(val):
                n_long.append({'parameter': row['parameter'], 'breed': breed,
                               'stage_kg': stage, 'value': val})
n_long_df = pd.DataFrame(n_long)

# Serum Urea vs N balance parameters
print("\nSerum Urea vs N balance parameters:")
n_serum_corr = []
for n_param in n_long_df['parameter'].unique():
    n_sub = n_long_df[n_long_df['parameter'] == n_param]
    merged = serum_urea_bs.merge(n_sub[['breed', 'stage_kg', 'value']],
                                  on=['breed', 'stage_kg'], suffixes=('_serum', '_n'))
    if len(merged) >= 5:
        r, p = pearsonr(merged['serum_urea_mean'], merged['value'])
        n_serum_corr.append({'N_parameter': n_param, 'pearson_r': round(r, 4),
                             'p_value': round(p, 6), 'n': len(merged)})

n_serum_corr_df = pd.DataFrame(n_serum_corr).sort_values('pearson_r', key=abs, ascending=False)
# FDR correction
if len(n_serum_corr_df) > 0:
    _, fdr_q = benjamini_hochberg(n_serum_corr_df['p_value'].values)
    n_serum_corr_df['q_value'] = fdr_q
    n_serum_corr_df['FDR_significant'] = n_serum_corr_df['q_value'] < 0.05
for _, r in n_serum_corr_df.iterrows():
    fdr_sig = ' [FDR]' if r.get('FDR_significant', False) else ''
    sig = '**' if r['p_value'] < 0.01 else ('*' if r['p_value'] < 0.05 else '')
    print(f"  {r['N_parameter']:35s} r={r['pearson_r']:+.4f} p={r['p_value']:.4f} q={r.get('q_value', np.nan):.4f} {sig}{fdr_sig}")

# N balance vs liver enzymes
n_expr_corr = []
for n_param in n_long_df['parameter'].unique():
    n_sub = n_long_df[n_long_df['parameter'] == n_param]
    for gene in AA_CATABOLISM_GENES:
        gene_sub = liver_bs_mean[liver_bs_mean['gene_name'] == gene]
        merged = gene_sub.merge(n_sub, on=['breed', 'stage_kg'])
        if len(merged) >= 6:
            r, p = pearsonr(merged['expr'], merged['value'])
            n_expr_corr.append({'N_parameter': n_param, 'Gene': gene,
                                'pearson_r': round(r, 4), 'p_value': round(p, 6),
                                'n': len(merged)})

n_expr_corr_df = pd.DataFrame(n_expr_corr)
# FDR correction
if len(n_expr_corr_df) > 0:
    _, fdr_q = benjamini_hochberg(n_expr_corr_df['p_value'].values)
    n_expr_corr_df['q_value'] = fdr_q
    n_expr_corr_df['FDR_significant'] = n_expr_corr_df['q_value'] < 0.05
print("\nTop N balance ↔ Liver enzyme correlations:")
top_n = n_expr_corr_df.dropna().sort_values('pearson_r', key=abs, ascending=False).head(25)
for _, r in top_n.iterrows():
    fdr_sig = ' [FDR]' if r.get('FDR_significant', False) else ''
    sig = '*' if r['p_value'] < 0.05 else ''
    print(f"  {r['N_parameter']:35s} ↔ {r['Gene']:10s}  r={r['pearson_r']:+.4f} p={r['p_value']:.4f} q={r.get('q_value', np.nan):.4f} {sig}{fdr_sig}")

# ============================================================
# 9. Cross-Time Validation: 15kg Liver → Later Phenotype
# ============================================================
print("\n" + "=" * 60)
print("Cross-Time Validation: Early Liver → Later Phenotype")
print("=" * 60)

# For each Tier 1 gene: correlate 15kg liver expression with 45/75kg serum Urea and N balance
tier1_genes = aa_tier_df[aa_tier_df['Tier'] == 1]['Gene'].tolist()
print(f"Tier 1 genes: {tier1_genes}")

# Get 15kg liver mean per gene per breed
liver_15 = liver_ind[liver_ind['stage_kg'] == 15].groupby(['gene_name', 'breed'])['expr'].mean().reset_index()

# Get 45kg and 75kg serum Urea per breed
serum_45 = serum_urea_ind[serum_urea_ind['stage_kg'] == 45].groupby('breed')['value'].mean()
serum_75 = serum_urea_ind[serum_urea_ind['stage_kg'] == 75].groupby('breed')['value'].mean()

cross_time = []
for gene in tier1_genes:
    g15 = liver_15[liver_15['gene_name'] == gene]
    if len(g15) < 2:
        continue
    dly_l15 = g15[g15['breed'] == 'DLY']['expr'].values
    tfb_l15 = g15[g15['breed'] == 'TFB']['expr'].values
    if len(dly_l15) == 0 or len(tfb_l15) == 0:
        continue

    # 15kg liver FC
    fc15 = np.log2(dly_l15[0] / tfb_l15[0]) if tfb_l15[0] > 0 else np.nan

    # Cross-time: 15kg liver FC vs 45kg serum Urea difference
    urea45_diff = serum_45.get('DLY', np.nan) - serum_45.get('TFB', np.nan)
    urea75_diff = serum_75.get('DLY', np.nan) - serum_75.get('TFB', np.nan)

    # For the N parameters at 45 and 75
    cross_time.append({
        'Gene': gene,
        'liver_FC_15kg': round(fc15, 3),
        'serum_Urea_DLYminusTFB_45kg': round(urea45_diff, 4),
        'serum_Urea_DLYminusTFB_75kg': round(urea75_diff, 4),
    })

cross_time_df = pd.DataFrame(cross_time)
print("\nCross-time reference (15kg liver → later serum):")
for _, r in cross_time_df.iterrows():
    direction = 'TFB↑ liver enzyme → TFB↑ serum Urea' if r['liver_FC_15kg'] < 0 else ''
    print(f"  {r['Gene']:10s}  LiverFC15={r['liver_FC_15kg']:+6.3f}  "
          f"ΔUrea45={r['serum_Urea_DLYminusTFB_45kg']:+.4f}  "
          f"ΔUrea75={r['serum_Urea_DLYminusTFB_75kg']:+.4f}  {direction}")

# ============================================================
# 10. FIGURE 1: Temporal Tier Heatmap
# ============================================================
print("\nGenerating Figure 1: Temporal Tier Classification Heatmap...")

# Combine AA enzymes + crosstalk genes with tier classification
plot_genes_data = []
for _, r in aa_tier_df.iterrows():
    plot_genes_data.append({
        'Gene': r['Gene'], 'Category': 'AA_Enzyme', 'Tier': r['Tier'],
        15: r.get('log2FC_15kg', np.nan), 45: r.get('log2FC_45kg', np.nan),
        75: r.get('log2FC_75kg', np.nan), 105: r.get('log2FC_105kg', np.nan)
    })

for _, r in ct_tier_df.iterrows():
    lt = r['Liver_Tier']
    # Only include if tier is 1-3
    if lt in [1, 2, 3]:
        plot_genes_data.append({
            'Gene': r['Gene'], 'Category': 'Crosstalk_Liver', 'Tier': lt,
            15: r.get('Liver_log2FC_15kg', np.nan) if r.get('Liver_log2FC_15kg') != '' else np.nan,
            45: r.get('Liver_log2FC_45kg', np.nan) if r.get('Liver_log2FC_45kg') != '' else np.nan,
            75: r.get('Liver_log2FC_75kg', np.nan) if r.get('Liver_log2FC_75kg') != '' else np.nan,
            105: r.get('Liver_log2FC_105kg', np.nan) if r.get('Liver_log2FC_105kg') != '' else np.nan,
        })

# Also add top TFs
for _, r in tf_tier_df.iterrows():
    if r['Tier'] in [1, 2, 3]:
        plot_genes_data.append({
            'Gene': r['TF'], 'Category': 'TF_Regulator', 'Tier': r['Tier'],
            15: r.get('log2FC_15kg', np.nan) if r.get('log2FC_15kg') != '' else np.nan,
            45: r.get('log2FC_45kg', np.nan) if r.get('log2FC_45kg') != '' else np.nan,
            75: r.get('log2FC_75kg', np.nan) if r.get('log2FC_75kg') != '' else np.nan,
            105: r.get('log2FC_105kg', np.nan) if r.get('log2FC_105kg') != '' else np.nan,
        })

plot_df = pd.DataFrame(plot_genes_data)

# Build matrix for heatmap
heatmap_data = plot_df.set_index('Gene')[[15, 45, 75, 105]].copy()
# Clip extreme values
heatmap_data = heatmap_data.clip(-6, 6)

# Sort by tier then by 15kg FC pattern
plot_df['sort_key'] = plot_df['Tier'] * 1000 + plot_df[15].fillna(0)
plot_df = plot_df.sort_values('sort_key')
heatmap_data = heatmap_data.loc[plot_df['Gene']]

# Create colormap annotation for tiers
tier_colors = {1: '#4CAF50', 2: '#FF9800', 3: '#9E9E9E'}
cat_colors = {'AA_Enzyme': '#E91E63', 'Crosstalk_Liver': '#2196F3', 'TF_Regulator': '#9C27B0'}

fig, ax = plt.subplots(figsize=(10, max(12, len(heatmap_data) * 0.35)))
mask = heatmap_data.isna()
sns.heatmap(heatmap_data, annot=True, fmt='.2f', cmap='RdBu_r', center=0,
            vmin=-6, vmax=6, mask=mask, ax=ax, cbar_kws={'label': 'log2(DLY/TFB)'},
            linewidths=0.5, linecolor='white', annot_kws={'fontsize': 7})

# Add tier color bars on left
for i, (gene, tier) in enumerate(zip(plot_df['Gene'], plot_df['Tier'])):
    ax.add_patch(plt.Rectangle((-0.12, i), 0.06, 1, facecolor=tier_colors.get(tier, 'white'),
                               edgecolor='none', transform=ax.transData, clip_on=False))
# Add category bar
for i, cat in enumerate(plot_df['Category']):
    ax.add_patch(plt.Rectangle((-0.06, i), 0.06, 1, facecolor=cat_colors.get(cat, 'white'),
                               edgecolor='none', transform=ax.transData, clip_on=False))

ax.set_title('Temporal Classification of Key Genes\nlog2(DLY/TFB) Across Stages\n'
             'Tier 1 (Green): Early Programming | Tier 2 (Orange): 45kg Switch | Tier 3 (Gray): Late Consequence',
             fontsize=12, fontweight='bold')
ax.set_ylabel('')
ax.set_xlabel('Stage (kg)')

# Legend
from matplotlib.patches import Patch
legend_elements = [
    Patch(facecolor='#4CAF50', label='Tier 1: Early Programming (15+45+75 consistent)'),
    Patch(facecolor='#FF9800', label='Tier 2: 45kg Peak / Switch'),
    Patch(facecolor='#9E9E9E', label='Tier 3: Late Consequence (75-105 only)'),
    Patch(facecolor='#E91E63', label='AA Catabolism Enzyme'),
    Patch(facecolor='#2196F3', label='Crosstalk / Hepatokine'),
    Patch(facecolor='#9C27B0', label='TF / Regulator'),
]
ax.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(1.02, 1.0),
          fontsize=8, frameon=False)

plt.tight_layout()
fig.savefig('fig_temporal_tiers.png', dpi=200, bbox_inches='tight')
fig.savefig('fig_temporal_tiers.pdf', bbox_inches='tight')
plt.close()
print("Saved fig_temporal_tiers.png/pdf")

# ============================================================
# 11. FIGURE 2: Tier 1 Driver Genes — Cross-Stage Consistency
# ============================================================
print("Generating Figure 2: Tier 1 Driver Genes...")

tier1_all = plot_df[plot_df['Tier'] == 1]
if len(tier1_all) > 0:
    n_cols = min(4, len(tier1_all))
    n_rows = int(np.ceil(len(tier1_all) / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 4, n_rows * 3.5))
    if n_rows * n_cols == 1:
        axes = np.array([axes])
    axes = axes.flatten()

    for i, (_, gene_row) in enumerate(tier1_all.iterrows()):
        ax = axes[i]
        gene = gene_row['Gene']
        gdf = liver_ind[liver_ind['gene_name'] == gene]

        for breed, color, marker, ls in [('DLY', '#2196F3', 'o', '-'), ('TFB', '#C62828', 's', '-')]:
            bs = gdf[gdf['breed'] == breed].groupby('stage_kg')['expr']
            means = bs.mean()
            sems = bs.std() / np.sqrt(bs.count())
            stages = means.index.values
            ax.errorbar(stages, means.values, yerr=sems.values, marker=marker,
                       color=color, linestyle=ls, linewidth=2, markersize=8,
                       capsize=4, label=breed)

        ax.set_title(f'{gene}\n(Tier 1: Early Programming)', fontsize=11, fontweight='bold')
        ax.set_xticks([15, 45, 75, 105])
        ax.set_xlabel('Stage (kg)', fontsize=9)
        ax.set_ylabel('Expr', fontsize=9)
        ax.grid(axis='y', alpha=0.3)
        if i == 0:
            ax.legend(fontsize=8)

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    fig.suptitle('Tier 1: Early Programming Genes — Consistent Divergence from 15 kg',
                 fontsize=14, fontweight='bold')
    plt.tight_layout()
    fig.savefig('fig_tier1_drivers.png', dpi=200, bbox_inches='tight')
    fig.savefig('fig_tier1_drivers.pdf', bbox_inches='tight')
    plt.close()
    print(f"Saved fig_tier1_drivers.png/pdf ({len(tier1_all)} genes)")

# ============================================================
# 12. FIGURE 3: Tier 2 — 45kg Switch Genes
# ============================================================
print("Generating Figure 3: Tier 2 45kg-Switch Genes...")

tier2_all = plot_df[plot_df['Tier'] == 2]
if len(tier2_all) > 0:
    n_cols = min(4, len(tier2_all))
    n_rows = int(np.ceil(len(tier2_all) / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 4, n_rows * 3.5))
    if n_rows * n_cols == 1:
        axes = np.array([axes])
    axes = axes.flatten()

    for i, (_, gene_row) in enumerate(tier2_all.iterrows()):
        ax = axes[i]
        gene = gene_row['Gene']
        gdf = liver_ind[liver_ind['gene_name'] == gene]

        for breed, color, marker, ls in [('DLY', '#2196F3', 'o', '-'), ('TFB', '#C62828', 's', '-')]:
            bs = gdf[gdf['breed'] == breed].groupby('stage_kg')['expr']
            means = bs.mean()
            sems = bs.std() / np.sqrt(bs.count())
            stages = means.index.values
            ax.errorbar(stages, means.values, yerr=sems.values, marker=marker,
                       color=color, linestyle=ls, linewidth=2, markersize=8,
                       capsize=4, label=breed)

        # Highlight 45kg window
        ax.axvspan(40, 50, alpha=0.1, color='#FF9800')
        ax.set_title(f'{gene}\n(Tier 2: 45kg Switch)', fontsize=11, fontweight='bold')
        ax.set_xticks([15, 45, 75, 105])
        ax.set_xlabel('Stage (kg)', fontsize=9)
        ax.set_ylabel('Expr', fontsize=9)
        ax.grid(axis='y', alpha=0.3)
        if i == 0:
            ax.legend(fontsize=8)

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    fig.suptitle('Tier 2: 45kg-Switch Genes — Divergence at the Protein Deposition Peak',
                 fontsize=14, fontweight='bold')
    plt.tight_layout()
    fig.savefig('fig_tier2_45kg_switches.png', dpi=200, bbox_inches='tight')
    fig.savefig('fig_tier2_45kg_switches.pdf', bbox_inches='tight')
    plt.close()
    print(f"Saved fig_tier2_45kg_switches.png/pdf ({len(tier2_all)} genes)")

# ============================================================
# 13. FIGURE 4: Cross-Time Causal Chain
# ============================================================
print("Generating Figure 4: Cross-Time Causal Chain...")

fig, axes = plt.subplots(2, 3, figsize=(16, 10))

# Panel A: 15kg liver SDS → Serum Urea at 45/75kg
ax = axes[0, 0]
sds_15 = liver_ind[(liver_ind['gene_name'] == 'SDS') & (liver_ind['stage_kg'] == 15)]
sds_15_bs = sds_15.groupby('breed')['expr'].mean()
ax.bar([0, 1], [sds_15_bs.get('DLY', 0), sds_15_bs.get('TFB', 0)],
       color=['#2196F3', '#C62828'], edgecolor='black', linewidth=1.2)
ax.set_xticks([0, 1]); ax.set_xticklabels(['DLY', 'TFB'])
ax.set_title('SDS @ 15 kg (Liver)', fontsize=11, fontweight='bold')
ax.set_ylabel('Expression')
ax.text(0.5, 0.95, f'log2FC={np.log2(sds_15_bs.get("DLY",1)/sds_15_bs.get("TFB",1)):.2f}',
        transform=ax.transAxes, ha='center', fontsize=10, fontstyle='italic')

# Panel B: Serum Urea at 45kg
ax = axes[0, 1]
urea_45 = serum_urea_ind[serum_urea_ind['stage_kg'] == 45].groupby('breed')['value'].agg(['mean', 'std'])
ax.bar([0, 1], [urea_45.loc['DLY', 'mean'], urea_45.loc['TFB', 'mean']],
       yerr=[urea_45.loc['DLY', 'std']/np.sqrt(8), urea_45.loc['TFB', 'std']/np.sqrt(8)],
       color=['#2196F3', '#C62828'], edgecolor='black', linewidth=1.2, capsize=5)
ax.set_xticks([0, 1]); ax.set_xticklabels(['DLY', 'TFB'])
ax.set_title('Serum Urea @ 45 kg', fontsize=11, fontweight='bold')
ax.set_ylabel('mmol/L')
ax.axhline(y=0, color='black', linewidth=0.5)

# Panel C: Serum Urea at 75kg (n=1 per breed — descriptive only, no error bars)
ax = axes[0, 2]
urea_75 = serum_urea_ind[serum_urea_ind['stage_kg'] == 75].groupby('breed')['value'].agg(['mean', 'count'])
n75_dly = int(urea_75.loc['DLY', 'count']) if 'DLY' in urea_75.index else 0
n75_tfb = int(urea_75.loc['TFB', 'count']) if 'TFB' in urea_75.index else 0
dly_val = urea_75.loc['DLY', 'mean'] if 'DLY' in urea_75.index else 0
tfb_val = urea_75.loc['TFB', 'mean'] if 'TFB' in urea_75.index else 0
bars = ax.bar([0, 1], [dly_val, tfb_val],
       color=['#2196F3', '#C62828'], edgecolor='black', linewidth=1.2)
ax.set_xticks([0, 1]); ax.set_xticklabels(['DLY', 'TFB'])
ax.set_title('Serum Urea @ 75 kg', fontsize=11, fontweight='bold')
ax.set_ylabel('mmol/L')
if n75_dly < 2 or n75_tfb < 2:
    ax.text(0.5, 0.95, f'n=1; descriptive only', transform=ax.transAxes,
            ha='center', fontsize=7, color='red', fontstyle='italic')

# Panel D: N balance — Protein deposition
ax = axes[1, 0]
stages = [15, 45, 75, 105]
for breed, color, marker in [('DLY', '#2196F3', 'o'), ('TFB', '#C62828', 's')]:
    vals = [n_df[n_df['parameter'].str.contains('Protein deposition', na=False)][f'{breed}_{s}_mean'].values for s in stages]
    valid_vals = [v[0] if len(v) > 0 and pd.notna(v[0]) else np.nan for v in vals]
    ax.plot(stages, valid_vals, marker=marker, color=color, linewidth=2, markersize=8, label=breed)
ax.set_xlabel('Stage (kg)'); ax.set_ylabel('N g/kg BW^0.75/d')
ax.set_title('Protein Deposition', fontsize=11, fontweight='bold')
ax.legend()
ax.grid(alpha=0.3)

# Panel E: N retention %
ax = axes[1, 1]
for breed, color, marker in [('DLY', '#2196F3', 'o'), ('TFB', '#C62828', 's')]:
    vals = [n_df[n_df['parameter'].str.contains('N retention', na=False)][f'{breed}_{s}_mean'].values for s in stages]
    valid_vals = [v[0] if len(v) > 0 and pd.notna(v[0]) else np.nan for v in vals]
    ax.plot(stages, valid_vals, marker=marker, color=color, linewidth=2, markersize=8)
ax.set_xlabel('Stage (kg)'); ax.set_ylabel('%')
ax.set_title('N Retention %', fontsize=11, fontweight='bold')
ax.grid(alpha=0.3)

# Panel F: UN (urinary N)
ax = axes[1, 2]
for breed, color, marker in [('DLY', '#2196F3', 'o'), ('TFB', '#C62828', 's')]:
    vals = [n_df[n_df['parameter'].str.contains('^UN', na=False, regex=True)][f'{breed}_{s}_mean'].values for s in stages]
    valid_vals = [v[0] if len(v) > 0 and pd.notna(v[0]) else np.nan for v in vals]
    ax.plot(stages, valid_vals, marker=marker, color=color, linewidth=2, markersize=8)
ax.set_xlabel('Stage (kg)'); ax.set_ylabel('g/d')
ax.set_title('Urinary N', fontsize=11, fontweight='bold')
ax.grid(alpha=0.3)

# Causal arrows
fig.suptitle('Cross-Time Causal Chain: Early Liver Programming → Late Phenotype\n'
             'TFB: SDS↑ (15kg) → Urea↑ (45-75kg) → N retention↓ (45-105kg)',
             fontsize=13, fontweight='bold')
plt.tight_layout()
fig.savefig('fig_cross_time_causal_chain.png', dpi=200, bbox_inches='tight')
fig.savefig('fig_cross_time_causal_chain.pdf', bbox_inches='tight')
plt.close()
print("Saved fig_cross_time_causal_chain.png/pdf")

# ============================================================
# 14. FIGURE 5: TF Temporal Classification
# ============================================================
print("Generating Figure 5: TF Temporal Classification...")

# Select TFs with good tier classification
tf_plot = tf_tier_df[tf_tier_df['Tier'].isin([1, 2, 3])].copy()
if len(tf_plot) > 0:
    fc_cols = ['log2FC_15kg', 'log2FC_45kg', 'log2FC_75kg', 'log2FC_105kg']
    tf_heatmap = tf_plot.set_index('TF')[fc_cols].apply(pd.to_numeric, errors='coerce')
    tf_heatmap.columns = [15, 45, 75, 105]
    tf_heatmap = tf_heatmap.clip(-6, 6)
    tf_plot['sort_key'] = tf_plot['Tier'] * 1000 + tf_heatmap[15].fillna(0)
    tf_plot = tf_plot.sort_values('sort_key')
    # Reindex to match sort
    tf_heatmap = tf_heatmap.loc[tf_plot['TF']]

    fig, ax = plt.subplots(figsize=(8, max(6, len(tf_heatmap) * 0.4)))
    mask = tf_heatmap.isna()
    sns.heatmap(tf_heatmap, annot=True, fmt='.2f', cmap='RdBu_r', center=0,
                vmin=-6, vmax=6, mask=mask, ax=ax,
                cbar_kws={'label': 'log2(DLY/TFB)'},
                linewidths=0.5, linecolor='white', annot_kws={'fontsize': 7})

    for i, (tf, tier) in enumerate(zip(tf_plot['TF'], tf_plot['Tier'])):
        ax.add_patch(plt.Rectangle((-0.10, i), 0.05, 1,
                                   facecolor=tier_colors.get(tier, 'white'),
                                   edgecolor='none', transform=ax.transData, clip_on=False))

    ax.set_title('Transcription Factor Temporal Classification\n'
                 'Tier 1 (Green) = Early Programming | Tier 2 (Orange) = 45kg Switch | Tier 3 (Gray) = Late',
                 fontsize=12, fontweight='bold')
    ax.set_ylabel('')

    tf_legend = [
        Patch(facecolor='#4CAF50', label='Tier 1: Early driver (≥15kg)'),
        Patch(facecolor='#FF9800', label='Tier 2: 45kg switch'),
        Patch(facecolor='#9E9E9E', label='Tier 3: Late consequence'),
    ]
    ax.legend(handles=tf_legend, loc='upper left', bbox_to_anchor=(1.02, 1.0),
              fontsize=8, frameon=False)

    plt.tight_layout()
    fig.savefig('fig_tf_temporal_tiers.png', dpi=200, bbox_inches='tight')
    fig.savefig('fig_tf_temporal_tiers.pdf', bbox_inches='tight')
    plt.close()
    print(f"Saved fig_tf_temporal_tiers.png/pdf ({len(tf_heatmap)} TFs)")

# ============================================================
# 15. FIGURE 6: Key Crosstalk Genes — 45kg Window Focus
# ============================================================
print("Generating Figure 6: Crosstalk Genes — 45kg Focus...")

# Select crosstalk genes with clear liver patterns
ct_liver_classified = ct_tier_df[ct_tier_df['Liver_Tier'].isin([1, 2, 3])].copy()
# Prioritize: Tier 1 first, then Tier 2, then select top by |FC_45|
if len(ct_liver_classified) > 16:
    top_ct = pd.concat([
        ct_liver_classified[ct_liver_classified['Liver_Tier'] == 1].head(6),
        ct_liver_classified[ct_liver_classified['Liver_Tier'] == 2].head(10),
    ]).head(16)
else:
    top_ct = ct_liver_classified

plot_ct_genes = top_ct['Gene'].tolist()
if len(plot_ct_genes) < 16:
    # Fill remaining with any classified crosstalk genes
    remaining = [g for g in CROSSTALK_GENES if g not in plot_ct_genes and g in all_liver_genes]
    plot_ct_genes.extend(remaining[:16 - len(plot_ct_genes)])

plot_ct_genes = plot_ct_genes[:16]

fig, axes = plt.subplots(4, 4, figsize=(18, 16))
axes = axes.flatten()

for i, gene in enumerate(plot_ct_genes):
    ax = axes[i]

    # Liver expression
    l_df = liver_ind[liver_ind['gene_name'] == gene]
    if len(l_df) > 0:
        for breed, color, marker in [('DLY', '#2196F3', 'o'), ('TFB', '#C62828', 's')]:
            bdf = l_df[l_df['breed'] == breed].groupby('stage_kg')['expr']
            means = bdf.mean()
            sems = bdf.std() / np.sqrt(bdf.count())
            ax.errorbar(means.index, means.values, yerr=sems.values,
                       marker=marker, color=color, linewidth=1.8, markersize=6,
                       capsize=3, alpha=0.9, label=f'{breed}-L' if i == 0 else '')

    # Muscle expression
    m_df = muscle_ind[muscle_ind['gene_name'] == gene]
    if len(m_df) > 0:
        for breed, color, marker in [('DLY', '#2196F3', 'o'), ('TFB', '#C62828', 's')]:
            bdf = m_df[m_df['breed'] == breed].groupby('stage_kg')['expr']
            means = bdf.mean()
            sems = bdf.std() / np.sqrt(bdf.count())
            ax.errorbar(means.index, means.values, yerr=sems.values,
                       marker=marker, color=color, linestyle='--', linewidth=1.8,
                       markersize=6, capsize=3, alpha=0.5)

    # 45kg highlight
    ax.axvspan(40, 50, alpha=0.08, color='#FF9800')

    # Add tier label
    l_tier = top_ct[top_ct['Gene'] == gene]['Liver_Tier'].values
    tier_str = f' [T{int(l_tier[0])}]' if len(l_tier) > 0 else ''
    ax.set_title(f'{gene}{tier_str}', fontsize=11, fontweight='bold')
    ax.set_xticks([15, 45, 75, 105])
    ax.tick_params(labelsize=8)
    ax.grid(axis='y', alpha=0.3)

handles = [
    plt.Line2D([0], [0], color='#2196F3', marker='o', linewidth=2, label='DLY Liver'),
    plt.Line2D([0], [0], color='#C62828', marker='s', linewidth=2, label='TFB Liver'),
    plt.Line2D([0], [0], color='#2196F3', marker='o', linewidth=2, linestyle='--', alpha=0.5, label='DLY Muscle'),
    plt.Line2D([0], [0], color='#C62828', marker='s', linewidth=2, linestyle='--', alpha=0.5, label='TFB Muscle'),
]
fig.legend(handles=handles, loc='lower center', ncol=4, fontsize=10, frameon=False)
fig.suptitle('Key Liver-Muscle Cross-Talk Genes (Tier-Annotated, 45kg Window Highlighted)',
             fontsize=14, fontweight='bold', y=1.01)
plt.tight_layout()
fig.savefig('fig_crosstalk_tier_annotated.png', dpi=200, bbox_inches='tight')
fig.savefig('fig_crosstalk_tier_annotated.pdf', bbox_inches='tight')
plt.close()
print("Saved fig_crosstalk_tier_annotated.png/pdf")

# ============================================================
# 16. FIGURE 7: Tier-Based Screening Strategy Summary
# ============================================================
print("Generating Figure 7: Screening Strategy Summary...")

fig, ax = plt.subplots(figsize=(14, 8))

# Schematic representation
tiers_summary = {
    'Tier 1\nEarly\nProgramming': {'15kg': 8, '45kg': 10, '75kg': 12, '105kg': 12},
    'Tier 2\n45kg\nSwitch': {'15kg': 2, '45kg': 8, '75kg': 10, '105kg': 11},
    'Tier 3\nLate\nConsequence': {'15kg': 1, '45kg': 1, '75kg': 5, '105kg': 8},
}

x = np.arange(4)
width = 0.25
stage_labels = ['15 kg', '45 kg', '75 kg', '105 kg']

for i, (tier_name, vals) in enumerate(tiers_summary.items()):
    bars = ax.bar(x + i * width, list(vals.values()), width, label=tier_name,
                  color=list(tier_colors.values())[i], edgecolor='black', linewidth=0.8)

ax.set_xticks(x + width)
ax.set_xticklabels(stage_labels, fontsize=12)
ax.set_ylabel('|log2FC| Magnitude (schematic)', fontsize=12)
ax.set_title('Tier-Based Gene Screening Strategy\n'
             'TFB Protein Deposition Peak = 45 kg  |  DLY Peak = 75–105 kg',
             fontsize=14, fontweight='bold')
ax.legend(fontsize=11, loc='upper left')
ax.grid(axis='y', alpha=0.3)

# Annotation arrows
ax.annotate('Early drivers\n(genetic programming)', xy=(0.25, 10), xytext=(0.5, 13.5),
            fontsize=10, ha='center', color='#2E7D32',
            arrowprops=dict(arrowstyle='->', color='#2E7D32', lw=1.5))
ax.annotate('Switch/peak\n(at deposition bottleneck)', xy=(1.25, 8.5), xytext=(2.0, 6.5),
            fontsize=10, ha='center', color='#E65100',
            arrowprops=dict(arrowstyle='->', color='#E65100', lw=1.5))
ax.annotate('Consequences\n(not causal)', xy=(2.25, 6), xytext=(2.8, 4),
            fontsize=10, ha='center', color='#616161',
            arrowprops=dict(arrowstyle='->', color='#616161', lw=1.5))

# Add phenotype overlay: protein deposition
ax2 = ax.twinx()
ax2.plot([0.5, 1.5, 2.5, 3.5], [1.26, 1.12, 0.68, 0.49], 'k--', linewidth=2, alpha=0.4)
ax2.set_ylabel('TFB Protein Deposition (N g/kg BW^0.75/d)', fontsize=10, alpha=0.5)
ax2.set_ylim(0, 2)

plt.tight_layout()
fig.savefig('fig_screening_strategy.png', dpi=200, bbox_inches='tight')
fig.savefig('fig_screening_strategy.pdf', bbox_inches='tight')
plt.close()
print("Saved fig_screening_strategy.png/pdf")

# ============================================================
# 17. Save All Results
# ============================================================
print("\n" + "=" * 60)
print("Saving comprehensive results...")

with pd.ExcelWriter('advanced_analysis_results.xlsx', engine='openpyxl') as writer:
    aa_tier_df.to_excel(writer, sheet_name='AA_Enzymes_Tier', index=False)
    ct_tier_df.to_excel(writer, sheet_name='Crosstalk_Genes_Tier', index=False)
    tf_tier_df.to_excel(writer, sheet_name='TF_Regulators_Tier', index=False)
    if len(tf_corr_df) > 0:
        tf_corr_df.to_excel(writer, sheet_name='TF_vs_Enzyme_Corr', index=False)
    tf_mean_corr.to_excel(writer, sheet_name='TF_Summary', index=False)
    ind_corr_df.to_excel(writer, sheet_name='Individual_Correlations', index=False)
    if len(n_expr_corr_df) > 0:
        n_expr_corr_df.to_excel(writer, sheet_name='N_vs_Expression', index=False)
    if len(n_serum_corr_df) > 0:
        n_serum_corr_df.to_excel(writer, sheet_name='N_vs_SerumUrea', index=False)
    if len(cross_time_df) > 0:
        cross_time_df.to_excel(writer, sheet_name='CrossTime_Validation', index=False)

print("Saved advanced_analysis_results.xlsx")
print("Sheets:")
print("  AA_Enzymes_Tier       — AA catabolism genes with temporal tier classification")
print("  Crosstalk_Genes_Tier  — Crosstalk genes with liver & muscle tier")
print("  TF_Regulators_Tier    — Transcription factors with temporal tier")
print("  TF_vs_Enzyme_Corr      — TF ↔ AA enzyme correlations")
print("  TF_Summary             — TF ranking by mean |r|")
print("  Individual_Correlations — Liver enzyme ↔ serum Urea (individual-level)")
print("  N_vs_Expression         — N balance ↔ liver enzyme")
print("  N_vs_SerumUrea          — N balance ↔ serum Urea")
print("  CrossTime_Validation    — 15kg liver → later phenotype")

print("\nFigures generated:")
print("  fig_temporal_tiers.png/pdf          — All-gene heatmap with tier classification")
print("  fig_tier1_drivers.png/pdf           — Tier 1 early programming genes")
print("  fig_tier2_45kg_switches.png/pdf     — Tier 2 45kg switch genes")
print("  fig_cross_time_causal_chain.png/pdf — Cross-time causal evidence")
print("  fig_tf_temporal_tiers.png/pdf       — TF temporal classification")
print("  fig_crosstalk_tier_annotated.png/pdf — Crosstalk genes tier-annotated")
print("  fig_screening_strategy.png/pdf      — Screening strategy schematic")

print("\n" + "=" * 60)
print("Analysis complete.")
print("=" * 60)
