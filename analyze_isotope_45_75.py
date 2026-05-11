#!/usr/bin/env python3
"""Corrected isotope tracer analysis for 45 kg and 75 kg using raw individual data.
Key fix: 75 kg col 0 = Q_protein (turnover in g protein/kg^0.75/d), NOT BW^0.75.
Q_gN = col0 / 6.25.
All P-values: Welch's independent t-test, two-tailed, computed from individual data.
"""

import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# =============================================================================
# Raw individual data — 75 kg
# Columns: [Q_protein, PS_gN, NS_abs_gN/d, PS_g_protein, PD_abs_gN/d?, PD_gN, PD_g_protein, dep_g_protein, dep_gN, ret_synth%]
# Q_gN = col[0] / 6.25
# =============================================================================

dl75_raw = {
    'DLY-2': [25.215, 3.383, 91.771, 21.145, 64.149, 2.365, 14.781, 6.365, 1.018, 30.10],
    'DLY-3': [29.905, 4.347, 121.161, 27.171, 89.709, 3.219, 20.118, 7.053, 1.128, 25.96],
    'DLY-4': [30.173, 4.199, 115.989, 26.244, 86.087, 3.117, 19.478, 6.766, 1.082, 25.78],
    'DLY-5': [27.447, 3.718, 101.790, 23.240, 74.562, 2.724, 17.024, 6.216, 0.995, 26.75],
    'DLY-6': [28.088, 4.069, 108.856, 25.434, 74.607, 2.789, 17.432, 8.002, 1.280, 31.46],
    'DLY-7': [21.347, 2.913, 82.259, 18.205, 49.513, 1.753, 10.958, 7.247, 1.160, 39.81],
}

tf75_raw = {
    'TFB-1': [20.42, 2.50, 66.72, 15.63, 49.45, 1.85, 11.59, 4.05, 0.65, 25.88],
    'TFB-2': [18.99, 2.17, 58.16, 13.58, 44.29, 1.65, 10.34, 3.24, 0.52, 23.85],
    'TFB-3': [16.40, 1.95, 50.42, 12.20, 31.62, 1.22,  7.65, 4.55, 0.73, 37.30],
    'TFB-5': [19.39, 2.44, 64.05, 15.27, 43.44, 1.66, 10.36, 4.91, 0.79, 32.17],
    'TFB-6': [17.87, 2.17, 57.81, 13.55, 39.47, 1.48,  9.75, 3.80, 0.69, 31.74],
    'TFB-7': [18.61, 2.25, 59.43, 14.04, 41.65, 1.57,  9.94, 4.11, 0.67, 30.19],
}

# =============================================================================
# Raw individual data — 45 kg
# Columns: [Q_gN, Turnover_protein, NS_gN, NS_abs_gN/d, PS_g_protein, D_abs, ND_gN, PD_g_protein, dep_g_protein, dep_N_g, ret_synth%]
# =============================================================================

d45_raw = {
    'DLY45-1': [5.576, 34.853, 5.03, 95.19,  31.41, 68.39,  3.61, 22.56, 8.84, 1.41, 28.16],
    'DLY45-2': [5.786, 36.165, 5.40, 103.88, 33.77, 71.33,  3.71, 23.19, 10.58, 1.69, 31.34],
    'DLY45-3': [5.303, 33.146, 4.86, 92.13,  30.40, 61.00,  3.22, 20.12, 10.27, 1.64, 33.79],
    'DLY45-4': [5.578, 34.863, 5.15, 101.23, 32.21, 70.54,  3.59, 22.45, 9.76, 1.56, 30.31],
    'DLY45-5': [7.393, 46.208, 7.08, 132.15, 44.26, 101.61, 5.44, 34.03, 10.23, 1.64, 23.11],
    'DLY45-6': [8.923, 55.770, 8.57, 164.81, 53.58, 130.75, 6.80, 42.51, 11.08, 1.77, 20.67],
    'DLY45-7': [4.429, 27.682, 3.92, 76.50,  24.51, 47.80,  2.45, 15.32, 9.20, 1.47, 37.51],
    'DLY45-8': [6.794, 42.460, 6.33, 119.04, 39.57, 90.05,  4.79, 29.93, 9.64, 1.54, 24.36],
}

t45_raw = {
    'TFB45-1': [6.790, 42.43,  6.12, 108.98, 38.25, 84.98,  4.77, 29.83, 8.42, 1.35, 22.02],
    'TFB45-2': [6.945, 43.41,  5.92, 103.74, 37.01, 87.30,  4.98, 31.14, 5.87, 0.94, 15.85],
    'TFB45-3': [6.641, 41.51,  5.95, 108.48, 37.18, 85.19,  4.67, 29.20, 7.98, 1.28, 21.48],
    'TFB45-4': [6.814, 42.59,  5.49, 100.55, 34.30, 87.82,  4.79, 29.96, 4.34, 0.69, 12.66],
    'TFB45-5': [7.198, 44.99,  6.20, 112.28, 38.72, 93.64,  5.17, 32.30, 6.43, 1.03, 16.60],
    'TFB45-6': [7.378, 46.11,  6.28, 112.27, 39.28, 94.11,  5.27, 32.93, 6.35, 1.02, 16.17],
    'TFB45-7': [7.624, 47.65,  6.87, 123.19, 42.96, 99.38,  5.55, 34.66, 8.30, 1.33, 19.33],
    'TFB45-8': [8.176, 51.10,  7.39, 129.95, 46.21, 106.91, 6.08, 38.01, 8.19, 1.31, 17.73],
}

# Parameter indices
# 45 kg: [Q_gN, Turnover_prot, PS_gN, NS_abs, PS_prot, D_abs, PD_gN, PD_prot, dep_prot, dep_N, ret_synth]
# 75 kg: [Q_prot, PS_gN, NS_abs, PS_prot, PD_abs?, PD_gN, PD_prot, dep_prot, dep_N, ret_synth]
param_45_idx = {'Q': 0, 'PS': 2, 'PD': 6, 'dep_N': 9, 'ret_synth': 10}
param_75_idx = {'Q': 0, 'PS': 1, 'PD': 5, 'dep_N': 8, 'ret_synth': 9}  # Q = col0/6.25

params = ['Q', 'PS', 'PD', 'dep_N', 'ret_synth']
param_labels_full = {
    'Q': 'N flux (Q), gN/kg BW^0.75/d',
    'PS': 'Protein synthesis (PS), gN/kg BW^0.75/d',
    'PD': 'Protein degradation (PD), gN/kg BW^0.75/d',
    'dep_N': 'Protein deposition, gN/kg BW^0.75/d',
    'ret_synth': 'Retention / Synthesis, %',
}
param_short = {
    'Q': 'N flux',
    'PS': 'Protein synthesis',
    'PD': 'Protein degradation',
    'dep_N': 'Protein deposition',
    'ret_synth': 'Retention/Synthesis',
}

# =============================================================================
# Extract individual-level parameter arrays
# =============================================================================

series = {}  # series[breed][stage][param] = np.array

for breed, stage, raw_dict, idx_map, q_div in [
    ('DLY', '45 kg', d45_raw, param_45_idx, 1.0),
    ('TFB', '45 kg', t45_raw, param_45_idx, 1.0),
    ('DLY', '75 kg', dl75_raw, param_75_idx, 6.25),
    ('TFB', '75 kg', tf75_raw, param_75_idx, 6.25),
]:
    key = (breed, stage)
    series[key] = {}
    for p in params:
        idx = idx_map[p]
        vals = np.array([v[idx] for v in raw_dict.values()], dtype=float)
        if p == 'Q' and q_div != 1.0:
            vals = vals / q_div  # Convert Q_protein to Q_gN
        series[key][p] = vals

# =============================================================================
# Descriptive stats
# =============================================================================

desc = {}  # desc[breed][stage][param] = {'n','mean','sd','sem'}
for (breed, stage), params_dict in series.items():
    key = (breed, stage)
    desc[key] = {}
    for p in params:
        v = params_dict[p]
        desc[key][p] = {
            'n': len(v),
            'mean': np.mean(v),
            'sd': np.std(v, ddof=1),
            'sem': np.std(v, ddof=1) / np.sqrt(len(v)),
        }

# =============================================================================
# Breed comparison at each stage: Welch's independent t-test, two-tailed
# =============================================================================

breed_tests = {}
for stage in ['45 kg', '75 kg']:
    breed_tests[stage] = {}
    for p in params:
        dly = series[('DLY', stage)][p]
        tfb = series[('TFB', stage)][p]
        t_stat, p_val = stats.ttest_ind(dly, tfb, equal_var=False)
        pooled_sd = np.sqrt((np.var(dly, ddof=1) + np.var(tfb, ddof=1)) / 2)
        d_cohen = (np.mean(dly) - np.mean(tfb)) / pooled_sd if pooled_sd > 0 else np.nan
        # Also check equal-variance for comparison
        _, p_equal = stats.ttest_ind(dly, tfb, equal_var=True)
        # Levene test
        _, p_levene = stats.levene(dly, tfb)
        breed_tests[stage][p] = {
            't_welch': t_stat, 'p_welch': p_val, 'p_equal': p_equal,
            'cohens_d': d_cohen, 'p_levene': p_levene,
        }

# =============================================================================
# Adjacent stage comparison (45 vs 75): Independent t-test
# =============================================================================

adj_tests = {}
for breed in ['DLY', 'TFB']:
    adj_tests[breed] = {}
    for p in params:
        v45 = series[(breed, '45 kg')][p]
        v75 = series[(breed, '75 kg')][p]
        t_stat, p_val = stats.ttest_ind(v45, v75, equal_var=False)
        adj_tests[breed][p] = {'t': t_stat, 'p': p_val}

# =============================================================================
# Print results
# =============================================================================

def sig(p):
    if np.isnan(p): return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return 'ns'

def fmt_p(p):
    return f'{p:.4f} {sig(p)}'

def fmt_ms(mean, sem, dec=2):
    return f'{mean:.{dec}f} ± {sem:.{dec}f}'

print("=" * 80)
print("CORRECTED ISOTOPE TRACER ANALYSIS — 45 kg & 75 kg")
print("Statistical method: Welch's independent t-test (two-tailed)")
print("=" * 80)

for stage in ['45 kg', '75 kg']:
    print(f"\n{'─'*70}")
    print(f"  {stage} — Descriptive Statistics (Mean ± SEM from individual data)")
    print(f"{'─'*70}")
    for p in params:
        for breed in ['DLY', 'TFB']:
            d = desc[(breed, stage)][p]
            print(f"  {breed:4s} {param_short[p]:30s}: {d['mean']:8.3f} ± {d['sem']:8.4f}  (n={d['n']}, SD={d['sd']:.3f})")

print(f"\n{'─'*70}")
print("  Breed Comparison: DLY vs TFB at Each Stage")
print(f"{'─'*70}")
for stage in ['45 kg', '75 kg']:
    print(f"\n  [{stage}]")
    for p in params:
        bt = breed_tests[stage][p]
        d_dly = desc[('DLY', stage)][p]
        d_tfb = desc[('TFB', stage)][p]
        var_ratio = (d_dly['sd']**2) / (d_tfb['sd']**2) if d_tfb['sd'] > 0 else np.inf
        print(f"  {param_short[p]:30s}: P_welch = {fmt_p(bt['p_welch']):16s}  "
              f"d = {bt['cohens_d']:.3f}  "
              f"Var_ratio = {var_ratio:.1f}  "
              f"(Levene P = {bt['p_levene']:.4f})")

print(f"\n{'─'*70}")
print("  Adjacent Stage Comparison: 45 vs 75 kg Within Each Breed")
print(f"{'─'*70}")
for breed in ['DLY', 'TFB']:
    print(f"\n  [{breed}]")
    for p in params:
        at = adj_tests[breed][p]
        d45 = desc[(breed, '45 kg')][p]
        d75 = desc[(breed, '75 kg')][p]
        print(f"  {param_short[p]:30s}: {fmt_ms(d45['mean'], d45['sem'])} → {fmt_ms(d75['mean'], d75['sem'])}  "
              f"P = {fmt_p(at['p']):16s}  (n={d45['n']},{d75['n']})")

# =============================================================================
# Compare with source file P-values
# =============================================================================

print(f"\n{'─'*70}")
print("  Comparison: Raw Data Recalculation vs Source Summary File")
print(f"{'─'*70}")

source_breed_p = {
    ('45 kg', 'Q'): 0.045, ('45 kg', 'PS'): 0.445, ('45 kg', 'PD'): 0.044,
    ('45 kg', 'dep_N'): 0.001, ('45 kg', 'ret_synth'): 0.014,
    ('75 kg', 'Q'): 0.001, ('75 kg', 'PS'): 0.001, ('75 kg', 'PD'): 0.001,
    ('75 kg', 'dep_N'): 0.001, ('75 kg', 'ret_synth'): 0.942,
}

print(f"\n  {'Parameter':30s} {'Stage':8s} {'P_raw (Welch 2-tail)':25s} {'P_source':15s} {'Agree?':10s} {'Note'}")
print(f"  {'─'*30} {'─'*8} {'─'*25} {'─'*15} {'─'*10} {'─'*30}")
for stage in ['45 kg', '75 kg']:
    for p in params:
        raw_p = breed_tests[stage][p]['p_welch']
        src_p = source_breed_p.get((stage, p), np.nan)
        # Agreement: both significant at 0.05 or both non-significant
        raw_sig = raw_p < 0.05
        src_sig = src_p < 0.05
        if np.isnan(src_p):
            agree = 'N/A'
            note = ''
        elif raw_sig == src_sig:
            agree = '✓'
            note = ''
        else:
            agree = '✗ DIFFER'
            note = f'Raw 2-tail P={raw_p:.4f}, source has P={src_p:.4f}'
            # Check if source might be one-tailed
            if abs(raw_p/2 - src_p) < 0.01:
                note += ' [source可能是单尾检验?]'
            else:
                note += ' [source可能用了等方差/不同检验]'
        print(f"  {param_short[p]:30s} {stage:8s} P = {raw_p:.4f} {sig(raw_p):4s}          P = {src_p:.4f} {sig(src_p):4s}  {agree:10s} {note}")

# =============================================================================
# Build Excel
# =============================================================================

wb = Workbook()

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUBHEADER_FONT = Font(name='Arial', size=10, bold=True)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=12, bold=True)
NOTE_FONT = Font(name='Arial', size=9, italic=True)
WARN_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_row(ws, row, ncols, bold_first=False):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT if (bold_first and c == 1) else BODY_FONT
        cell.alignment = CENTER if c > 1 else LEFT
        cell.border = THIN_BORDER

def auto_width(ws, min_w=10, max_w=30):
    for col_cells in ws.columns:
        cl = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                lengths.append(w)
        if lengths:
            ws.column_dimensions[cl].width = min(max(max(lengths)+2, min_w), max_w)

# ---- Sheet 1: Individual Data ----

ws1 = wb.active
ws1.title = '1_Raw_Individual_Data'

ws1.merge_cells('A1:L1')
ws1.cell(row=1, column=1, value='Isotope Tracer Individual Data — 45 kg & 75 kg').font = TITLE_FONT

r = 3
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
ws1.cell(row=r, column=1, value='45 kg (DLY = 1-1~1-8, TFB = 2-1~2-8)').font = SUBHEADER_FONT
r += 1
h45 = ['ID', 'Breed', 'Q (gN)', 'Turnover (prot)', 'PS (gN)', 'NS_abs (gN/d)',
       'PS (prot)', 'D_abs', 'PD (gN)', 'PD (prot)', 'Dep (prot)', 'Dep (gN)', 'Ret/Synth%']
for ci, h in enumerate(h45, 1):
    ws1.cell(row=r, column=ci, value=h)
style_header(ws1, r, len(h45))
r += 1

for breed_label, raw_dict in [('DLY', d45_raw), ('TFB', t45_raw)]:
    for pig_id, vals in raw_dict.items():
        ws1.cell(row=r, column=1, value=pig_id)
        ws1.cell(row=r, column=2, value=breed_label)
        for j, v in enumerate(vals):
            ws1.cell(row=r, column=3+j, value=round(v, 3))
        style_row(ws1, r, len(h45))
        r += 1

r += 2
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
ws1.cell(row=r, column=1, value='75 kg (DLY-2~7, TFB-1~7 except TFB-4)').font = SUBHEADER_FONT
r += 1
h75 = ['ID', 'Breed', 'Q_prot (g prot)', 'PS (gN)', 'NS_abs (gN/d)',
       'PS (prot)', 'PD_abs?', 'PD (gN)', 'PD (prot)', 'Dep (prot)', 'Dep (gN)', 'Ret/Synth%', 'Q_gN (calc)']
for ci, h in enumerate(h75, 1):
    ws1.cell(row=r, column=ci, value=h)
style_header(ws1, r, len(h75))
r += 1

for breed_label, raw_dict in [('DLY', dl75_raw), ('TFB', tf75_raw)]:
    for pig_id, vals in raw_dict.items():
        ws1.cell(row=r, column=1, value=pig_id)
        ws1.cell(row=r, column=2, value=breed_label)
        for j in range(len(vals)):
            ws1.cell(row=r, column=3+j, value=round(vals[j], 3))
        ws1.cell(row=r, column=15, value=round(vals[0]/6.25, 3))  # Q_gN
        style_row(ws1, r, len(h75))
        r += 1

auto_width(ws1)

# ---- Sheet 2: Descriptive Statistics ----

ws2 = wb.create_sheet('2_Descriptive_Mean_SEM')

ws2.merge_cells('A1:F1')
ws2.cell(row=1, column=1, value='Isotope Tracer — Descriptive Statistics (Mean ± SEM from Individual Data)').font = TITLE_FONT

ncols2 = 6
headers2 = ['Parameter', 'DLY 45 kg', 'TFB 45 kg', 'DLY 75 kg', 'TFB 75 kg', 'Note']
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=ci, value=h)
style_header(ws2, 3, ncols2)

for i, p in enumerate(params):
    r = 4 + i
    ws2.cell(row=r, column=1, value=param_labels_full[p])
    for j, key in enumerate([('DLY','45 kg'), ('TFB','45 kg'), ('DLY','75 kg'), ('TFB','75 kg')]):
        d = desc[key][p]
        ws2.cell(row=r, column=2+j, value=fmt_ms(d['mean'], d['sem'], dec=2))
    ws2.cell(row=r, column=6, value='Computed from raw individual data')
    style_row(ws2, r, ncols2)

r_n = 4 + len(params)
ws2.cell(row=r_n, column=1, value='n')
for j, key in enumerate([('DLY','45 kg'), ('TFB','45 kg'), ('DLY','75 kg'), ('TFB','75 kg')]):
    ws2.cell(row=r_n, column=2+j, value=desc[key]['PS']['n'])
style_row(ws2, r_n, ncols2, bold_first=True)

r_note = r_n + 2
ws2.merge_cells(start_row=r_note, start_column=1, end_row=r_note, end_column=6)
ws2.cell(row=r_note, column=1,
         value='Note: Values are Mean ± SEM computed from individual animal data. '
         '75 kg Q_gN calculated as Q_protein(col 0) / 6.25. '
         'Statistical test: Welch\'s independent t-test (two-tailed), not assuming equal variance.').font = NOTE_FONT
auto_width(ws2)

# ---- Sheet 3: Breed Comparison P-values ----

ws3 = wb.create_sheet('3_Breed_Comparison_P')

ws3.merge_cells('A1:J1')
ws3.cell(row=1, column=1,
         value='Breed Comparison (DLY vs TFB) — Welch\'s Independent t-test, Two-tailed').font = TITLE_FONT

ncols3 = 10
headers3 = ['Parameter', 'Stage', 'DLY Mean±SEM', 'TFB Mean±SEM', 'P_welch (2-tail)',
            'Sig', "Cohen's d", 'P_equal_var', 'Levene P', 'Note']
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=ci, value=h)
style_header(ws3, 3, ncols3)

r = 4
for stage in ['45 kg', '75 kg']:
    for p in params:
        bt = breed_tests[stage][p]
        d_dly = desc[('DLY', stage)][p]
        d_tfb = desc[('TFB', stage)][p]
        ws3.cell(row=r, column=1, value=param_short[p])
        ws3.cell(row=r, column=2, value=stage)
        ws3.cell(row=r, column=3, value=fmt_ms(d_dly['mean'], d_dly['sem']))
        ws3.cell(row=r, column=4, value=fmt_ms(d_tfb['mean'], d_tfb['sem']))
        ws3.cell(row=r, column=5, value=round(bt['p_welch'], 4))
        ws3.cell(row=r, column=6, value=sig(bt['p_welch']))
        ws3.cell(row=r, column=7, value=round(bt['cohens_d'], 3))
        ws3.cell(row=r, column=8, value=round(bt['p_equal'], 4))
        ws3.cell(row=r, column=9, value=round(bt['p_levene'], 4))
        # Flag if variance ratio > 3
        var_ratio = (d_dly['sd']**2) / (d_tfb['sd']**2) if d_tfb['sd'] > 0 else np.inf
        note = ''
        if var_ratio > 3:
            note = f'Var ratio={var_ratio:.1f}x; Welch needed'
        elif var_ratio < 0.33:
            note = f'Var ratio={var_ratio:.1f}x; Welch needed'
        ws3.cell(row=r, column=10, value=note)
        # Highlight rows with large variance discrepancy
        if var_ratio > 3 or var_ratio < 0.33:
            for c in range(1, ncols3+1):
                ws3.cell(row=r, column=c).fill = WARN_FILL
        style_row(ws3, r, ncols3)
        r += 1
auto_width(ws3)

# ---- Sheet 4: Adjacent Stage P-values ----

ws4 = wb.create_sheet('4_Adjacent_Stage_P')

ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1,
         value='Adjacent Stage Comparison (45 vs 75 kg) — Independent t-test (Different Pigs at Different Stages)').font = TITLE_FONT

ncols4 = 9
headers4 = ['Parameter', 'Breed', '45 kg Mean±SEM', '75 kg Mean±SEM', 'P-value', 'Sig', 'Test', 'Note']
for ci, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=ci, value=h)
style_header(ws4, 3, ncols4)

r = 4
for breed in ['DLY', 'TFB']:
    for p in params:
        at = adj_tests[breed][p]
        d45 = desc[(breed, '45 kg')][p]
        d75 = desc[(breed, '75 kg')][p]
        ws4.cell(row=r, column=1, value=param_short[p])
        ws4.cell(row=r, column=2, value=breed)
        ws4.cell(row=r, column=3, value=fmt_ms(d45['mean'], d45['sem']))
        ws4.cell(row=r, column=4, value=fmt_ms(d75['mean'], d75['sem']))
        ws4.cell(row=r, column=5, value=round(at['p'], 4))
        ws4.cell(row=r, column=6, value=sig(at['p']))
        ws4.cell(row=r, column=7, value='Independent t')
        ws4.cell(row=r, column=8, value='Different pigs at 45 vs 75 kg')
        style_row(ws4, r, ncols4)
        r += 1
auto_width(ws4)

# ---- Sheet 5: Source Comparison ----

ws5 = wb.create_sheet('5_Source_vs_Raw_Comparison')

ws5.merge_cells('A1:J1')
ws5.cell(row=1, column=1,
         value='Comparison: Raw Individual Data Recalculation vs Source Summary File P-values').font = TITLE_FONT

ncols5 = 10
headers5 = ['Parameter', 'Stage', 'P_raw (Welch 2-tail)', 'Sig', 'P_source', 'Src_Sig',
            'Same conclusion?', 'Raw means', 'Source means', 'Diagnosis']
for ci, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=ci, value=h)
style_header(ws5, 3, ncols5)

source_means = {
    ('DLY','45 kg'): {'Q':6.22,'PS':5.79,'PD':4.20,'dep_N':1.59,'ret_synth':28.66},
    ('TFB','45 kg'): {'Q':7.20,'PS':6.28,'PD':5.16,'dep_N':1.12,'ret_synth':17.73},
    ('DLY','75 kg'): {'Q':4.32,'PS':3.77,'PD':2.66,'dep_N':1.11,'ret_synth':29.98},
    ('TFB','75 kg'): {'Q':2.98,'PS':2.25,'PD':1.57,'dep_N':0.68,'ret_synth':30.19},
}

r = 4
for stage in ['45 kg', '75 kg']:
    for p in params:
        raw_p = breed_tests[stage][p]['p_welch']
        src_p = source_breed_p.get((stage, p), np.nan)
        raw_sig = raw_p < 0.05
        src_sig = src_p < 0.05 if not np.isnan(src_p) else None
        if src_sig is None:
            same = 'N/A'
            diag = ''
        elif raw_sig == src_sig:
            same = 'YES'
            diag = ''
        else:
            same = 'NO ← DIFFER'
            diag = f'Two-tailed Welch P={raw_p:.4f} vs source P={src_p:.4f}'
            if abs(raw_p/2 - src_p) < 0.015:
                diag += ' | Source likely used ONE-TAILED test (P_two_tail/2 ≈ P_source)'
            else:
                diag += ' | Possible equal-variance t-test or different data'

        dly_src_m = source_means[('DLY',stage)][p]
        tfb_src_m = source_means[('TFB',stage)][p]
        dly_raw_m = desc[('DLY',stage)][p]['mean']
        tfb_raw_m = desc[('TFB',stage)][p]['mean']

        ws5.cell(row=r, column=1, value=param_short[p])
        ws5.cell(row=r, column=2, value=stage)
        ws5.cell(row=r, column=3, value=round(raw_p, 4))
        ws5.cell(row=r, column=4, value=sig(raw_p))
        ws5.cell(row=r, column=5, value=src_p)
        ws5.cell(row=r, column=6, value=sig(src_p) if not np.isnan(src_p) else '')
        ws5.cell(row=r, column=7, value=same)
        ws5.cell(row=r, column=8, value=f'DLY={dly_raw_m:.2f}, TFB={tfb_raw_m:.2f}')
        ws5.cell(row=r, column=9, value=f'DLY={dly_src_m:.2f}, TFB={tfb_src_m:.2f}')
        ws5.cell(row=r, column=10, value=diag)
        style_row(ws5, r, ncols5)
        if same.startswith('NO'):
            for c in range(1, ncols5+1):
                ws5.cell(row=r, column=c).fill = WARN_FILL
        r += 1

# Add note
r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws5.cell(row=r, column=1,
         value='Key finding: At 45 kg, source file P-values for Q and PD appear to use ONE-TAILED tests '
         '(P_two_tail/2 ≈ P_source). Proper two-tailed Welch test shows NO significant breed difference '
         'for N flux and protein degradation at 45 kg. DLY variance is 7-9x larger than TFB at 45 kg, '
         'invalidating equal-variance assumptions.').font = NOTE_FONT
auto_width(ws5)

# =============================================================================
# Save
# =============================================================================

output_path = 'isotope_45vs75_corrected.xlsx'
wb.save(output_path)
print(f"\n{'='*60}")
print(f'Saved: {output_path}')
print(f'Sheets: {wb.sheetnames}')
